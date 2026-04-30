"""Client-facing weekly schedule: template copy into a workbook + placeholder fill."""

from __future__ import annotations

from collections import Counter
from copy import copy as shallow_copy
from dataclasses import fields
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from .calculate import WEEKS_PER_MONTH, CalibratedSchedule, round_half_up, schedule_to_dict
from .extract import ExtractedForm
from .validate import ValidationReport
from .templates.weekly_schedule_tokens import (
    WEEKLY_SCHEDULE_TOKENS,
    WEEK_SCHEDULE_DAY_CODES,
)

_TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "weekly_schedule_template.xlsx"

FULL_DOW: tuple[str, ...] = (
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
)

# Template columns are Sunday-first; keys match ``WEEK_SCHEDULE_DAY_CODES``.
_CODE_TO_FULL: dict[str, str] = {
    "SUN": "Sunday",
    "MON": "Monday",
    "TUE": "Tuesday",
    "WED": "Wednesday",
    "THU": "Thursday",
    "FRI": "Friday",
    "SAT": "Saturday",
}


def format_hrs_mins(minutes: int) -> str:
    """Format minutes as ``\"H hrs MM mins\"`` (two-digit minutes, non-negative)."""
    m = int(minutes)
    if m < 0:
        m = 0
    h, mm = divmod(m, 60)
    return f"{h} hrs {mm:02d} mins"


def _fmt_variance_mins(delta: int) -> str:
    if delta == 0:
        return "0 mins"
    if delta > 0:
        return f"{delta} mins"
    return f"-{abs(delta)} mins"


def _schedule_dict(schedule: CalibratedSchedule | dict[str, Any]) -> dict[str, Any]:
    if isinstance(schedule, CalibratedSchedule):
        return schedule_to_dict(schedule)
    if isinstance(schedule, dict):
        return schedule
    raise TypeError(
        f"schedule must be CalibratedSchedule or dict, got {type(schedule).__name__}"
    )


def _weekly_pattern(sd: dict[str, Any]) -> dict[str, Any]:
    return dict(sd.get("weekly_pattern") or sd.get("days") or {})


def _shift_start_default(wp: dict[str, Any]) -> str:
    starts: list[str] = []
    for dow in FULL_DOW:
        s = str((wp.get(dow) or {}).get("start") or "").strip()
        if s:
            starts.append(s)
    if not starts:
        return ""
    return Counter(starts).most_common(1)[0][0]


def _copy_sheet_into(
    src_ws: Worksheet,
    dest_wb: WorkbookType,
    *,
    sheet_title: str = "Weekly Schedule",
    insert_at: int | None = 1,
) -> Worksheet:
    """Duplicate *src_ws* into *dest_wb* with styles (openpyxl has no merged copy API).

    Copies cell values/styles, merges, dimensions, freeze panes, and grid-line visibility.
    """
    dest_ws = dest_wb.create_sheet(title=sheet_title, index=insert_at)

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    if cell.font:
                        new_cell.font = shallow_copy(cell.font)
                    if cell.fill:
                        new_cell.fill = shallow_copy(cell.fill)
                    if cell.border:
                        new_cell.border = shallow_copy(cell.border)
                    if cell.alignment:
                        new_cell.alignment = shallow_copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                    if cell.protection:
                        new_cell.protection = shallow_copy(cell.protection)
                except (AttributeError, TypeError):
                    pass

    for mr in list(src_ws.merged_cells.ranges):
        dest_ws.merge_cells(str(mr))

    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width is not None:
            dest_ws.column_dimensions[col_letter].width = col_dim.width

    for row_idx, row_dim in src_ws.row_dimensions.items():
        if row_dim.height is not None:
            dest_ws.row_dimensions[row_idx].height = row_dim.height

    dest_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dest_ws.freeze_panes = src_ws.freeze_panes
    return dest_ws


def _apply_placeholder_substitutions(ws: Worksheet, pairs: list[tuple[str, str]]) -> None:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            new_v = v
            for ph, rep in pairs:
                if ph in new_v:
                    new_v = new_v.replace(ph, rep)
            cell.value = new_v


def _placeholder_replacements(
    extracted_form: ExtractedForm,
    sd: dict[str, Any],
    *,
    shift_start_override: str | None,
) -> list[tuple[str, str]]:
    """``(placeholder, value)`` pairs, longest placeholder first for safe ``str.replace``."""
    wp = _weekly_pattern(sd)
    tok = WEEKLY_SCHEDULE_TOKENS

    shift_start = (shift_start_override or "").strip() or _shift_start_default(wp)

    vals: dict[str, str] = {
        "CLIENT_NAME": extracted_form.client_name or "",
        "PROVIDER_NAME": extracted_form.provider_name or "",
        "SHIFT_START": shift_start,
    }

    for code in WEEK_SCHEDULE_DAY_CODES:
        full = _CODE_TO_FULL[code]
        day = wp.get(full) or {}
        vals[f"CLOCK_IN_{code}"] = str(day.get("start") or "")
        vals[f"CLOCK_OUT_{code}"] = str(day.get("end") or "")
        mins = int(day.get("minutes") or 0)
        vals[f"TOTAL_{code}"] = format_hrs_mins(mins)
        tasks = list(day.get("tasks") or [])
        for i in range(12):
            key = f"TASK_{code}_{i + 1}"
            if i < len(tasks):
                vals[key] = f"{i + 1}. {tasks[i]}"
            else:
                vals[key] = ""

    weekly_total_min = sum(int((wp.get(dow) or {}).get("minutes") or 0) for dow in FULL_DOW)
    monthly_proj = round_half_up(weekly_total_min * WEEKS_PER_MONTH)
    authorized = int(sd.get("mdhhs_monthly_minutes") or 0)
    delivered_min = int(sd.get("delivered_minutes") or 0)
    variance = monthly_proj - authorized

    vals["WEEKLY_TOTAL"] = format_hrs_mins(weekly_total_min)
    vals["MONTHLY_TOTAL"] = format_hrs_mins(monthly_proj)
    vals["MONTHLY_DELIVERED"] = format_hrs_mins(delivered_min)
    vals["AUTHORIZED_TOTAL"] = format_hrs_mins(authorized)
    vals["VARIANCE"] = _fmt_variance_mins(variance)
    vals["STATUS"] = (
        "EXACT MATCH ✓" if variance == 0 else f"CHECK — {variance} min variance"
    )

    pairs: list[tuple[str, str]] = []
    for f in fields(tok):
        ph = getattr(tok, f.name)
        pairs.append((ph, vals[f.name]))
    pairs.sort(key=lambda x: len(x[0]), reverse=True)
    return pairs


def _find_weekly_status_cell(ws: Worksheet):
    """Return the cell that holds the narrative status line after template fill."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if "EXACT MATCH" in v or s.startswith("CHECK"):
                return cell
    return None


def _finalize_weekly_schedule_sheet(
    dest_ws: Worksheet,
    sd: dict[str, Any],
    validation_report: ValidationReport | None,
) -> None:
    dest_ws.sheet_view.showGridLines = False

    wp = _weekly_pattern(sd)
    weekly_total_min = sum(int((wp.get(dow) or {}).get("minutes") or 0) for dow in FULL_DOW)
    monthly_proj = round_half_up(weekly_total_min * WEEKS_PER_MONTH)
    delivered = int(sd.get("delivered_minutes") or 0)

    badge_labels = {
        "BILLABLE_EXACT": "BILLABLE EXACT ✓",
        "BILLABLE_AT_CAP": "BILLABLE AT CAP ✓",
        "BILLABLE_UNDER_CAP": "BILLABLE UNDER CAP ✓",
        "INVALID": "INVALID ✗",
    }
    cal_txt = ""
    if delivered != monthly_proj:
        delta = delivered - monthly_proj
        ad = abs(delta)
        if delta > 0:
            cal_txt = f"CALENDAR OVERSHOOT +{ad} min"
        else:
            cal_txt = f"CALENDAR SHORT −{ad} min"

    badge = ""
    if validation_report is not None:
        vstat = getattr(validation_report, "validation_status", "INVALID")
        badge = badge_labels.get(vstat, str(vstat))

    parts = [x for x in (cal_txt, badge) if x]
    cell = _find_weekly_status_cell(dest_ws)
    if parts and cell is not None:
        cell.value = " · ".join(parts)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    dest_ws.print_options.horizontalCentered = True
    ori = getattr(dest_ws, "ORIENTATION_LANDSCAPE", "landscape")
    dest_ws.page_setup.orientation = ori
    dest_ws.page_setup.fitToHeight = 0
    dest_ws.page_setup.fitToWidth = 1
    try:
        from openpyxl.worksheet.properties import PageSetupProperties

        sp = dest_ws.sheet_properties
        if sp is not None:
            sp.pageSetUpPr = PageSetupProperties(fitToPage=True)
    except (ImportError, TypeError, AttributeError):
        pass
    dest_ws.print_title_rows = "1:3"


def build_weekly_schedule_sheet(
    extracted_form: ExtractedForm,
    schedule: CalibratedSchedule | dict[str, Any],
    wb: WorkbookType,
    *,
    shift_start_override: str | None = None,
    insert_at: int | None = 1,
    sheet_title: str = "Weekly Schedule",
    validation_report: ValidationReport | None = None,
) -> None:
    """Copy ``weekly_schedule_template.xlsx``'s weekly sheet into *wb*, then substitute tokens."""
    sd = _schedule_dict(schedule)
    pairs = _placeholder_replacements(
        extracted_form, sd, shift_start_override=shift_start_override
    )

    if not _TEMPLATE_PATH.is_file():
        msg = f"Weekly schedule template missing: {_TEMPLATE_PATH}"
        raise FileNotFoundError(msg)

    template_wb = load_workbook(_TEMPLATE_PATH, keep_vba=False)
    try:
        src_ws = template_wb["Weekly Schedule"]
        dest_ws = _copy_sheet_into(
            src_ws, wb, sheet_title=sheet_title, insert_at=insert_at
        )
        _apply_placeholder_substitutions(dest_ws, pairs)
        _finalize_weekly_schedule_sheet(dest_ws, sd, validation_report)
    finally:
        template_wb.close()
