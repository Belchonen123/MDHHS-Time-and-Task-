"""Placeholder tokens for the client-facing weekly schedule workbook.

Values match the `Field Guide` sheet in `weekly_schedule_template.xlsx`.
Renderers should import :data:`WEEKLY_SCHEDULE_TOKENS` (or field names from
:class:`WeeklyScheduleTokens`) instead of embedding ``{{...}}`` literals.
"""

from __future__ import annotations

from dataclasses import field, fields, make_dataclass
from pathlib import Path
from typing import Final

WEEK_SCHEDULE_DAY_CODES: Final[tuple[str, ...]] = (
    "SUN",
    "MON",
    "TUE",
    "WED",
    "THU",
    "FRI",
    "SAT",
)
_DAY_HEADERS: Final[tuple[str, ...]] = (
    "Sunday",
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
)


def _placeholder(token_name: str) -> str:
    return "{{" + token_name + "}}"


def _weekly_schedule_token_field_specs() -> list[tuple[str, str, str]]:
    """Return (field_name, token_name, description) for the Field Guide."""
    rows: list[tuple[str, str, str]] = [
        ("CLIENT_NAME", "CLIENT_NAME", "Client display name in the schedule title."),
        ("SHIFT_START", "SHIFT_START", "Authorized shift start time (header subtext)."),
        ("PROVIDER_NAME", "PROVIDER_NAME", "Provider name (header subtext)."),
    ]
    for d in WEEK_SCHEDULE_DAY_CODES:
        rows.append((f"CLOCK_IN_{d}", f"CLOCK_IN_{d}", f"Clock-in time for {d}."))
        rows.append((f"CLOCK_OUT_{d}", f"CLOCK_OUT_{d}", f"Clock-out time for {d}."))
        rows.append((f"TOTAL_{d}", f"TOTAL_{d}", f"Total hours worked for {d}."))
    for d in WEEK_SCHEDULE_DAY_CODES:
        for i in range(1, 13):
            rows.append(
                (
                    f"TASK_{d}_{i}",
                    f"TASK_{d}_{i}",
                    f"Task line {i} for {d} (client tasks grid).",
                )
            )
    for name, desc in (
        ("WEEKLY_TOTAL", "Sum of daily hours for the week."),
        ("MONTHLY_TOTAL", "Monthly projection (e.g. weekly × 4.3)."),
        (
            "MONTHLY_DELIVERED",
            "Calendar-month delivered minutes (Σ daily schedule).",
        ),
        ("AUTHORIZED_TOTAL", "MDHHS-6064-P authorized hours."),
        ("VARIANCE", "Difference vs authorized."),
        ("STATUS", "Narrative status for the variance."),
    ):
        rows.append((name, name, desc))
    return rows


def _make_weekly_schedule_tokens_class() -> type:
    field_defs = [
        (field_name, str, field(default=_placeholder(token_name)))
        for field_name, token_name, _desc in _weekly_schedule_token_field_specs()
    ]
    return make_dataclass(
        "WeeklyScheduleTokens",
        field_defs,
        frozen=True,
        slots=True,
    )


WeeklyScheduleTokens = _make_weekly_schedule_tokens_class()
WEEKLY_SCHEDULE_TOKENS: Final[WeeklyScheduleTokens] = WeeklyScheduleTokens()


def write_weekly_schedule_template(dest: Path | None = None) -> Path:
    """Build `weekly_schedule_template.xlsx` next to this module (for maintenance)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    t = WEEKLY_SCHEDULE_TOKENS
    path = dest or Path(__file__).with_name("weekly_schedule_template.xlsx")

    font_name = "Arial"
    thin = Side(style="thin", color="FF808080")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Weekly Schedule"

    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_name, color="FFFFFF", bold=True, size=11)
    label_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

    # --- Title band (merged) — matches client-facing spec -----------------
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"WEEKLY SCHEDULE — {t.CLIENT_NAME}"
    title_cell.font = Font(name=font_name, size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = (
        f"Based on MDHHS-6064-P authorization • Shift start: {t.SHIFT_START} • "
        f"Provider: {t.PROVIDER_NAME}"
    )
    sub.font = Font(name=font_name, size=10, italic=True, color="FF595959")
    sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    ws.row_dimensions[3].height = 6

    # Clock grid
    clock_top = 4
    corner = ws.cell(row=clock_top, column=1, value="Weekly Schedule")
    corner.font = Font(name=font_name, bold=True, size=10)
    corner.fill = label_fill
    corner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    corner.border = grid_border
    for col, label in enumerate(_DAY_HEADERS, start=2):
        c = ws.cell(row=clock_top, column=col, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = grid_border

    clock_labels = ("Clock In", "Clock Out", "Total Time Worked")
    clock_attrs = (
        [f"CLOCK_IN_{d}" for d in WEEK_SCHEDULE_DAY_CODES],
        [f"CLOCK_OUT_{d}" for d in WEEK_SCHEDULE_DAY_CODES],
        [f"TOTAL_{d}" for d in WEEK_SCHEDULE_DAY_CODES],
    )
    for r_off, label, attr_names in zip(range(3), clock_labels, clock_attrs, strict=True):
        row = clock_top + 1 + r_off
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = label_fill
        lc.font = Font(name=font_name, bold=True, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = grid_border
        for col, attr in enumerate(attr_names, start=2):
            c = ws.cell(row=row, column=col, value=getattr(t, attr))
            c.font = Font(name=font_name, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = grid_border

    task_header_row = clock_top + 5
    th_corner = ws.cell(row=task_header_row, column=1, value="Client Tasks:")
    th_corner.font = Font(name=font_name, bold=True, size=10)
    th_corner.fill = label_fill
    th_corner.alignment = Alignment(horizontal="left", vertical="center")
    th_corner.border = grid_border
    for col, label in enumerate(_DAY_HEADERS, start=2):
        c = ws.cell(row=task_header_row, column=col, value=f"{label} Tasks")
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = grid_border

    first_task_row = task_header_row + 1
    for slot in range(12):
        row = first_task_row + slot
        idx = ws.cell(row=row, column=1, value=slot + 1)
        idx.font = Font(name=font_name, size=10)
        idx.alignment = Alignment(horizontal="center", vertical="top")
        idx.border = grid_border
        for col, day in enumerate(WEEK_SCHEDULE_DAY_CODES, start=2):
            attr = f"TASK_{day}_{slot + 1}"
            c = ws.cell(row=row, column=col, value=getattr(t, attr))
            c.font = Font(name=font_name, size=10)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = grid_border

    summary_row = first_task_row + 13
    pairs = (
        ("Weekly Total Time:", t.WEEKLY_TOTAL),
        ("Monthly Projected (x 4.3 weeks):", t.MONTHLY_TOTAL),
        ("Monthly Delivered (calendar shape):", t.MONTHLY_DELIVERED),
        ("MDHHS Authorized (per 6064-P):", t.AUTHORIZED_TOTAL),
        ("Variance:", t.VARIANCE),
        ("Status:", t.STATUS),
    )
    for i, (lbl, ph) in enumerate(pairs):
        r = summary_row + i
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name=font_name, bold=True, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        lc.border = grid_border
        vc = ws.cell(row=r, column=2, value=ph)
        vc.font = Font(name=font_name, size=10, bold=i in (1, 2))
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc.border = grid_border

    # Column widths — wide enough for full task names (e.g. Shopping for Food/Meds)
    ws.column_dimensions["A"].width = 30
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = ws.cell(row=first_task_row, column=2)

    fg = wb.create_sheet("Field Guide")
    fg.append(["Token", "Placeholder", "Description"])
    for hdr in fg[1]:
        hdr.fill = header_fill
        hdr.font = header_font
    specs = _weekly_schedule_token_field_specs()
    for f, (_fname, token_name, desc) in zip(
        fields(WEEKLY_SCHEDULE_TOKENS), specs, strict=True
    ):
        if f.name != _fname:
            msg = f"Field Guide spec order mismatch: {f.name} vs {_fname}"
            raise ValueError(msg)
        fg.append([token_name, getattr(WEEKLY_SCHEDULE_TOKENS, f.name), desc])

    wb.save(path)
    return path


if __name__ == "__main__":
    write_weekly_schedule_template()
