"""Render the plan-of-care workbook (7 tabs) from a month-calibrated schedule.

Hard rule — the xlsx builder is **rendering and math only**. It reads the
month-aware schedule exactly as ``calculate.generate_schedule`` produced it
and never infers, regenerates, or second-guesses placement. The single
sources of truth are:

* ``schedule["daily_schedule"]`` — one entry per scheduled day with
  ``date``, ``day_of_week``, ``shift_type``, ``clock_in/out``,
  ``duration_min``, and ``tasks`` (task-name → minutes).
* ``schedule["task_occurrence_counts"]`` — per-task session counts that
  match the daily list; used by Task Reconciliation so the workbook
  doesn't re-derive counts from a hardcoded frequency table.

No legacy day-of-week placement helpers, no "Wed = housework day"
colour rules, no hardcoded WEEKDAY_STD/HW_DAY/WEEKEND_FULL/CATCHUP row
labels — every per-day or per-shift breakdown aggregates from
``daily_schedule`` at render time.

Produces tabs (in order):
    1. Summary
    2. Weekly Schedule (client-facing 7-day grid from template)
    3. Weekly Pattern
    4. Daily Schedule        <- invariant: Σ task_minutes == duration_min
    5. Schedule Math
    6. Task Reconciliation
    7. Instructions

After saving we best-effort call LibreOffice headless to recalc formulas so
consumers using ``data_only=True`` get cached values. Missing LibreOffice is
a soft failure (log warning, keep openpyxl output).
"""

from __future__ import annotations

import logging
import re
import shutil
import statistics
import subprocess
import tempfile
from datetime import date, time
from decimal import ROUND_HALF_UP, ROUND_HALF_EVEN, Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

try:
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.series import SeriesLabel
except ImportError:  # pragma: no cover — optional chart support
    LineChart = None
    Reference = None
    GraphicalProperties = None
    LineProperties = None
    SeriesLabel = None

from .calculate import (
    WEEKS_PER_MONTH,
    CalibratedSchedule,
    compute_mdhhs_form_amount,
    compute_monthly_minutes_rounded,
    compute_task_amount,
    round_half_up,
    schedule_to_dict,
)
from .build_weekly_schedule import build_weekly_schedule_sheet
from .extract import ExtractedForm
from .validate import (
    EHHS_THRESHOLD_MIN,
    TRAVEL_TASKS_EXCLUDED_FROM_EHHS,
    ValidationReport,
)

logger = logging.getLogger(__name__)


def _money_from_minutes(minutes: int, pay_rate: float) -> float:
    """MDHHS-style money rounding — half-up to the cent from minutes × pay / 60."""
    return float(
        Decimal(str(int(minutes) * float(pay_rate) / 60.0)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    )


# --- Color & font palette (per spec) -----------------------------------------

FONT_NAME = "Arial"

COLOR_INPUT = "0000FF"          # hardcoded MDHHS auth values / task names
COLOR_FORMULA = "000000"        # formulas & computed values
COLOR_XREF = "008000"           # cross-tab references
COLOR_WHITE = "FFFFFF"

FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")            # yellow
FILL_SUBTOTAL_EHHS = PatternFill(
    "solid", fgColor="FFF9E8"
)  # pale yellow — EHHS-bound subtotal (muted TOTAL)
FILL_SUBTOTAL_TRAVEL = PatternFill(
    "solid", fgColor="FFF3E8"
)  # slightly warmer muted — travel subtotal
FILL_DEVIATION = PatternFill("solid", fgColor="FFE5CC")        # orange (catchup)
FILL_SECTION = PatternFill("solid", fgColor="1F4E78")          # dark blue
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")           # medium blue
FILL_WEEKEND = PatternFill("solid", fgColor="FCE4D6")          # peach
FILL_NONE = PatternFill()
FILL_LIGHT = PatternFill("solid", fgColor="DEEBF7")            # pale blue for labels
FILL_WEEKBAND = PatternFill("solid", fgColor="EAEFF7")         # ice blue — weekly subtotal band
FILL_EHHS_OK = PatternFill("solid", fgColor="C6EFCE")         # pale green — under EHHS cap
FILL_EHHS_EXCEED = PatternFill("solid", fgColor="FFC7CE")     # pale red — EHHS approval required

COLOR_EHHS_OK_FONT = "006100"
COLOR_EHHS_EXCEED_FONT = "9C0000"

_THIN = Side(style="thin", color="808080")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# --- Layout primitives -------------------------------------------------------
# Standard column widths / row heights so tabs share visual rhythm.
LAYOUT: dict[str, float | int] = {
    "label_col_width": 26,
    "value_col_width": 18,
    "value_wide_width": 24,
    "secondary_col_width": 14,
    "wide_text_width": 56,
    "task_min_col_width": 12,
    "section_row_height": 22,
    "header_row_height": 18,
    "data_row_height": 18,
    "tall_data_row_height": 26,
    "title_row_height": 28,
}

# Very pale variance hint (per-task rows, Task Reconciliation)
FILL_VARIANCE_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_VARIANCE_NEG = PatternFill("solid", fgColor="FFC7CE")
FILL_TITLE_ACCENT = PatternFill("solid", fgColor="17365D")
FILL_TABLE_BLUE = "TableStyleMedium2"
FILL_TABLE_GREEN = "TableStyleMedium4"

TAB_COLORS: dict[str, str] = {
    "Summary": "17365D",
    "Weekly Schedule": "1F4E78",
    "Weekly Pattern": "2E75B6",
    "Daily Schedule": "5B9BD5",
    "Schedule Math": "70AD47",
    "Task Reconciliation": "FFC000",
    "Instructions": "A5A5A5",
}


def _apply_block_layout(
    ws: Worksheet,
    *,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    fill: PatternFill | None = None,
) -> None:
    """Borders (+ optional fill) across a contiguous rectangle."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if fill is not None and fill is not FILL_NONE:
                cell.fill = fill


def _spacer_row(ws: Worksheet, row: int, *, height: int = 8) -> None:
    ws.row_dimensions[row].height = float(height)


def _kpi_tile(
    ws: Worksheet,
    *,
    row_start: int,
    col_start: int,
    col_end: int,
    label: str,
    value: float | int | str,
    footer: str,
    bg_hex: str,
    value_fmt: str | None = None,
) -> None:
    """Three-row KPI tile: tiny label row, big value row, tiny footer row."""
    white = "FFFFFF"
    bg = PatternFill("solid", fgColor=bg_hex)

    segments: tuple[tuple[str, Font, float, str | None], ...] = (
        (
            label,
            Font(name=FONT_NAME, size=9, bold=True, color=white),
            18,
            None,
        ),
        (
            value,
            Font(name=FONT_NAME, size=22, bold=True, color=white),
            40,
            value_fmt,
        ),
        (
            footer,
            Font(name=FONT_NAME, size=9, italic=True, color=white),
            16,
            None,
        ),
    )
    for i, (text, font, rh, nf) in enumerate(segments):
        r = row_start + i
        ws.merge_cells(
            start_row=r,
            start_column=col_start,
            end_row=r,
            end_column=col_end,
        )
        c = ws.cell(r, col_start)
        c.value = text
        c.font = font
        c.fill = bg
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if nf:
            c.number_format = nf
        ws.row_dimensions[r].height = rh


def _set_print_layout(
    ws: Worksheet,
    *,
    landscape: bool = True,
    fit_to_width: bool = True,
    print_title_rows: str | None = "1:3",
) -> None:
    """Printing: centered, fitted to width, repeat top rows."""
    ws.print_options.horizontalCentered = True
    ori = getattr(ws, "ORIENTATION_LANDSCAPE", "landscape")
    orp = getattr(ws, "ORIENTATION_PORTRAIT", "portrait")
    ws.page_setup.orientation = ori if landscape else orp
    if fit_to_width:
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        try:
            from openpyxl.worksheet.properties import PageSetupProperties

            sp = ws.sheet_properties
            if sp is not None:
                sp.pageSetUpPr = PageSetupProperties(fitToPage=True)
        except (ImportError, TypeError, AttributeError):
            pass
    if print_title_rows:
        ws.print_title_rows = print_title_rows


def _f_body() -> Font:
    return Font(name=FONT_NAME, size=10, color=COLOR_FORMULA)


def _f_input() -> Font:
    return Font(name=FONT_NAME, size=10, color=COLOR_INPUT)


def _f_xref() -> Font:
    return Font(name=FONT_NAME, size=10, color=COLOR_XREF)


def _f_col_header() -> Font:
    return Font(name=FONT_NAME, size=11, color=COLOR_WHITE, bold=True)


def _f_section() -> Font:
    return Font(name=FONT_NAME, size=14, color=COLOR_WHITE, bold=True)


def _f_sub_header() -> Font:
    return Font(name=FONT_NAME, size=10, color=COLOR_FORMULA, bold=True)


def _f_total() -> Font:
    return Font(name=FONT_NAME, size=10, color=COLOR_FORMULA, bold=True)


def _f_total_muted(*, italic: bool = True) -> Font:
    return Font(
        name=FONT_NAME,
        size=10,
        color=COLOR_FORMULA,
        bold=True,
        italic=italic,
    )


def _fmt_hours_from_minutes(minutes: int) -> str:
    """Whole minutes → ``\"174.92 hr\"`` (decimal hours, matches EHHS wording)."""
    h = round(int(minutes) / 60.0, 2)
    return f"{h:.2f} hr"


SHORT_DOW = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")
FULL_DOW = (
    "Monday", "Tuesday", "Wednesday", "Thursday",
    "Friday", "Saturday", "Sunday",
)


# --- Helpers -----------------------------------------------------------------


def _as_schedule_dict(schedule: Any) -> dict[str, Any]:
    if isinstance(schedule, CalibratedSchedule):
        return schedule_to_dict(schedule)
    if isinstance(schedule, dict):
        return schedule
    raise TypeError(
        f"schedule must be CalibratedSchedule or dict, got {type(schedule).__name__}"
    )


def _parse_iso(d: str) -> date | None:
    try:
        return date.fromisoformat(d)
    except (TypeError, ValueError):
        return None


def _ampm_to_time(s: str) -> time | str:
    t = (s or "").strip()
    m = re.fullmatch(r"(\d{1,2}):(\d{2})\s*([AP]M)", t, re.IGNORECASE)
    if not m:
        return t
    h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    h24 = (0 if h == 12 else h) if ap == "AM" else (12 if h == 12 else h + 12)
    return time(h24, mi)


def _f_muted() -> Font:
    """Muted text for secondary raw-minute columns (Reviewer reads HH:MM first)."""
    return Font(name=FONT_NAME, size=10, color="808080")


def _fmt_hhmm(minutes: int | None) -> str:
    """``HH:MM`` (zero-padded), matching the printed MDHHS-6064-P.

    Negative durations get a leading ``-`` (used in variance columns).
    Empty / None / 0 returns empty string so blank cells stay blank.
    """
    if minutes is None:
        return ""
    m = int(minutes)
    if m == 0:
        return ""
    sign = "-" if m < 0 else ""
    h, r = divmod(abs(m), 60)
    return f"{sign}{h:02d}:{r:02d}"


def _fmt_hm_words(total_min: int | None) -> str:
    """Whole minutes → readable ``44h 26m`` / ``2h 5m`` / ``45m`` (Plan of Care totals)."""
    if total_min is None:
        return ""
    m = int(total_min)
    if m == 0:
        return ""
    sign = "-" if m < 0 else ""
    abs_m = abs(m)
    h, mm = divmod(abs_m, 60)
    if h and mm:
        body = f"{h}h {mm}m"
    elif h:
        body = f"{h}h"
    else:
        body = f"{mm}m"
    return f"{sign}{body}"


def _fmt_hhmm_with_words(minutes: int | None) -> str:
    """``HH:MM`` plus ``(44h 26m)`` when both forms are non-empty — for spreadsheet totals."""
    hm = _fmt_hhmm(minutes)
    words = _fmt_hm_words(minutes)
    if hm and words:
        return f"{hm} ({words})"
    return hm or words


def _fmt_hours_minutes_long(minutes: int | None) -> str:
    """Readable workbook display: ``40 hrs 30 mins`` with zero-padded minutes."""
    if minutes is None:
        return ""
    m = int(minutes)
    if m == 0:
        return "0 hrs 00 mins"
    sign = "-" if m < 0 else ""
    h, mm = divmod(abs(m), 60)
    return f"{sign}{h} hrs {mm:02d} mins"


def _safe_table_name(name: str) -> str:
    """Excel table names: letters/numbers/underscore, unique enough for this workbook."""
    clean = re.sub(r"[^A-Za-z0-9_]", "_", name.strip())
    if not clean or not clean[0].isalpha():
        clean = f"T_{clean}"
    return clean[:240]


def _add_excel_table(
    ws: Worksheet,
    *,
    ref: str,
    name: str,
    style: str = FILL_TABLE_BLUE,
) -> None:
    """Best-effort real Excel table for filters/banding; formatting still works if skipped."""
    try:
        tab = Table(displayName=_safe_table_name(name), ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
    except ValueError:
        logger.debug("Skipping duplicate/overlapping Excel table %s on %s", name, ws.title)


def _professionalize_sheet(
    ws: Worksheet,
    *,
    title_rows: int = 1,
    freeze: str | None = None,
    auto_filter: str | None = None,
) -> None:
    """Global finishing pass: tab colors, page margins, view, filters, and title polish."""
    if ws.title in TAB_COLORS:
        ws.sheet_properties.tabColor = TAB_COLORS[ws.title]
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    if freeze:
        ws.freeze_panes = freeze
    if auto_filter:
        ws.auto_filter.ref = auto_filter
    try:
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.55
        ws.page_margins.bottom = 0.45
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2
    except AttributeError:
        pass
    for r in range(1, min(title_rows, ws.max_row) + 1):
        ws.row_dimensions[r].height = max(float(ws.row_dimensions[r].height or 0), 24.0)


def _ordered_task_names(
    extracted_form: ExtractedForm, daily: list[dict[str, Any]]
) -> list[str]:
    """Column order for Daily Schedule: extracted-form order first, then any
    extra names seen only in the schedule."""
    order: list[str] = []
    seen: set[str] = set()
    for t in extracted_form.tasks:
        nm = str(t.get("task_name") or "").strip()
        if nm and nm not in seen:
            seen.add(nm)
            order.append(nm)
    for d in daily:
        for nm in (d.get("tasks") or {}).keys():
            if nm and nm not in seen:
                seen.add(nm)
                order.append(nm)
    return order


def _day_fill(dow_short: str, is_catchup: bool) -> PatternFill:
    """Visual fill for a row in the Daily Schedule / Weekly Pattern.

    Only two cues — weekend rows (Sat/Sun) and the catch-up day if one
    exists on this plan. Deliberately generic: no weekday-specific
    fills (e.g. "Wed = housework day") because the scheduler can place
    any task on any day.
    """
    if is_catchup:
        return FILL_DEVIATION
    if dow_short in ("Sat", "Sun"):
        return FILL_WEEKEND
    return FILL_NONE


def _apply_border_range(ws: Worksheet, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = BORDER


def _set_col_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _title_row(
    ws: Worksheet,
    text: str,
    row: int,
    col_start: int,
    col_end: int,
    *,
    height: float | None = None,
) -> None:
    ws.merge_cells(
        start_row=row, start_column=col_start, end_row=row, end_column=col_end
    )
    c = ws.cell(row, col_start, value=text)
    c.font = _f_section()
    c.fill = FILL_SECTION
    c.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(col_start, col_end + 1):
        ws.cell(row, cc).fill = FILL_SECTION
        ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = float(height or LAYOUT["title_row_height"])


def _sub_title(
    ws: Worksheet, text: str, row: int, col_start: int, col_end: int,
    fill: PatternFill | None = None,
) -> None:
    ws.merge_cells(
        start_row=row, start_column=col_start, end_row=row, end_column=col_end
    )
    c = ws.cell(row, col_start, value=text)
    c.font = _f_col_header()
    c.fill = fill or FILL_HEADER
    c.alignment = Alignment(horizontal="left", vertical="center")
    for cc in range(col_start, col_end + 1):
        ws.cell(row, cc).fill = fill or FILL_HEADER
        ws.cell(row, cc).border = BORDER
    ws.row_dimensions[row].height = float(LAYOUT["section_row_height"])


def _col_headers(ws: Worksheet, row: int, headers: list[str], col0: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row, col0 + i, value=h)
        c.font = _f_col_header()
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = float(LAYOUT["header_row_height"])


def _shift_type_stats(
    daily: list[dict[str, Any]],
    catchup_iso: str | None = None,
) -> list[tuple[str, int, int, bool]]:
    """Aggregate (shift_type, days, total_min, is_catchup) from daily_schedule.

    Derived at render time so the workbook has no knowledge of which
    shift-type names are "expected" — whatever the scheduler emits is
    what gets displayed. The ``is_catchup`` flag is true when every
    day of that shift type falls on the catchup date (normally a
    singleton CATCHUP row); it drives the orange highlight only.
    """
    # Preserve first-seen order so output is deterministic for tests.
    order: list[str] = []
    counts: dict[str, int] = {}
    mins: dict[str, int] = {}
    dates_by_shift: dict[str, list[str]] = {}
    for d in daily:
        st = str(d.get("shift_type") or "").strip()
        if not st:
            continue
        if st not in counts:
            order.append(st)
            counts[st] = 0
            mins[st] = 0
            dates_by_shift[st] = []
        counts[st] += 1
        mins[st] += int(d.get("duration_min") or 0)
        dates_by_shift[st].append(str(d.get("date") or ""))
    catchup = str(catchup_iso) if catchup_iso else ""
    out: list[tuple[str, int, int, bool]] = []
    for st in order:
        is_catch = bool(catchup) and dates_by_shift[st] == [catchup]
        out.append((st, counts[st], mins[st], is_catch))
    return out


def _authorized_non_travel_travel_totals(form: ExtractedForm) -> tuple[int, int]:
    """Sum ``compute_monthly_minutes_rounded`` by EHHS-bound vs excluded travel."""
    non_travel = 0
    travel_only = 0
    for t in form.tasks:
        nm = str(t.get("task_name") or "")
        mpd = int(t.get("min_per_day") or 0)
        dpw = int(t.get("days_per_week") or 0)
        auth_min = compute_monthly_minutes_rounded(mpd, dpw)
        if nm in TRAVEL_TASKS_EXCLUDED_FROM_EHHS:
            travel_only += auth_min
        else:
            non_travel += auth_min
    return non_travel, travel_only


def _fmt_calendar_overshoot_dollars(amount: float) -> str:
    """Display variance dollars as ``+$20.25`` / ``−$12.34`` (sign outside ``$``)."""
    d = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN)
    if d == 0:
        return "$0.00"
    s = f"{abs(float(d)):,.2f}"
    typographic_minus = "\u2212"
    if d > 0:
        return f"+${s}"
    return f"{typographic_minus}${s}"


# --- Tab 1: Summary ----------------------------------------------------------


def _build_summary(
    ws: Worksheet,
    form: ExtractedForm,
    sd: dict[str, Any],
    vr: ValidationReport,
) -> None:
    ws.sheet_view.showGridLines = False
    dr = float(LAYOUT["data_row_height"])

    month_name = str(sd.get("month_name") or "")
    year = int(sd.get("year") or 0)

    _title_row(
        ws,
        f"Plan of Care Summary — {form.client_name or 'Client'} — {month_name} {year}",
        row=1,
        col_start=1,
        col_end=6,
    )

    daily = list(sd.get("daily_schedule") or [])
    catchup_iso = sd.get("catchup_date")
    pay_sr = float(sd.get("pay_rate") or 0.0)
    auth_min = int(sd.get("mdhhs_monthly_minutes") or 0)
    auth_amt = float(
        sd.get("mdhhs_form_amount") or sd.get("mdhhs_monthly_amount") or 0.0
    )

    raw_del = sd.get("delivered_minutes")
    if raw_del is None:
        del_min = sum(int(d.get("duration_min") or 0) for d in daily)
    else:
        del_min = int(raw_del)

    raw_del_amt = sd.get("delivered_amount")
    if raw_del_amt is None:
        raw_d = Decimal(del_min) * Decimal(str(pay_sr)) / Decimal("60")
        del_amt = float(raw_d.quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN))
    else:
        del_amt = float(raw_del_amt)

    raw_bm = sd.get("billable_minutes")
    bill_min = min(del_min, auth_min) if raw_bm is None else int(raw_bm)

    raw_ba = sd.get("billable_amount")
    if raw_ba is None:
        db = Decimal(str(del_amt)).quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN)
        ab = Decimal(str(auth_amt)).quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN)
        bill_amt = float(min(db, ab))
    else:
        bill_amt = float(raw_ba)

    ov_min = del_min - auth_min
    ov_amt = float(
        (Decimal(str(del_amt)) - Decimal(str(auth_amt))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_EVEN
        )
    )
    if ov_min > 0:
        tail_label = "Calendar overshoot (non-billable)"
        tail_note = "normal in some 31-day months"
    elif ov_min < 0:
        tail_label = "Under-delivery (bill actual)"
        tail_note = "normal in 28-day Feb or 4× weekday months"
    else:
        tail_label = "Calendar variance"
        tail_note = (
            "Authorization caps billing; delivered follows the calendar; "
            "billable = min(delivered, authorized) per ASM 144."
        )

    badge_labels = {
        "BILLABLE_EXACT": "BILLABLE EXACT ✓",
        "BILLABLE_AT_CAP": "BILLABLE AT CAP ✓",
        "BILLABLE_UNDER_CAP": "BILLABLE UNDER CAP ✓",
        "INVALID": "INVALID ✗",
    }
    vstat = getattr(vr, "validation_status", "INVALID")
    badge = badge_labels.get(vstat, vstat)

    _spacer_row(ws, 2, height=10)

    KPI_NAVY = "1F3864"
    KPI_BLUE = "2E5C9E"
    KPI_EMERALD = "1B5E2F"
    KPI_BRONZE = "8C5A1A"
    MONEY_FMT = '"$"#,##0.00;[Red]"($"#,##0.00")"'
    MONEY_SIGN = '"+$"#,##0.00;[Red]"-$"#,##0.00;"$0.00"'

    hh_b = _fmt_hhmm(bill_min) or ""
    bill_hm_long = _fmt_hours_minutes_long(bill_min)
    footer_bill = f"{bill_min:,} min" + (f"  ·  {hh_b}" if hh_b else "")
    ah = _fmt_hhmm(auth_min) or ""
    auth_hm_long = _fmt_hours_minutes_long(auth_min)
    footer_auth = f"{auth_min:,} min" + (f"  ·  {ah}" if ah else "")

    dd_h = _fmt_hhmm(del_min) or ""
    del_hm_long = _fmt_hours_minutes_long(del_min)
    del_footer = f"{del_min:,} min" + (f"  ·  {dd_h}" if dd_h else "")

    footer_bill = f"{footer_bill} / {bill_hm_long}"
    footer_auth = f"{footer_auth} / {auth_hm_long}"
    del_footer = f"{del_footer} / {del_hm_long}"

    _kpi_tile(
        ws,
        row_start=3,
        col_start=1,
        col_end=2,
        label="AUTHORIZED CAP",
        value=auth_amt,
        footer=footer_auth,
        bg_hex=KPI_NAVY,
        value_fmt=MONEY_FMT,
    )
    _kpi_tile(
        ws,
        row_start=3,
        col_start=3,
        col_end=4,
        label="DELIVERED (CALENDAR)",
        value=del_amt,
        footer=del_footer,
        bg_hex=KPI_BLUE,
        value_fmt=MONEY_FMT,
    )
    _kpi_tile(
        ws,
        row_start=3,
        col_start=5,
        col_end=5,
        label="BILLABLE (PER ASM 144)",
        value=bill_amt,
        footer=footer_bill,
        bg_hex=KPI_EMERALD,
        value_fmt=MONEY_FMT,
    )
    _kpi_tile(
        ws,
        row_start=3,
        col_start=6,
        col_end=6,
        label="CALENDAR OVERSHOOT",
        value=ov_amt,
        footer=f"{ov_min:+d} min",
        bg_hex=KPI_BRONZE,
        value_fmt=MONEY_SIGN,
    )

    _spacer_row(ws, 6, height=10)

    _sub_title(ws, "CLIENT INFORMATION", 7, 1, 3)
    _sub_title(ws, "ADULT SERVICES WORKER", 7, 4, 6)

    client_rows: list[tuple[str, Any]] = [
        ("Client Name", form.client_name),
        ("Client ID", form.client_id),
        ("County", form.county_name),
        ("Case Number", form.case_number),
        ("Provider Agency", form.provider_name),
        ("Pay Rate ($/hr)", float(form.pay_rate or 0.0)),
        ("Auth Date", form.auth_date),
    ]
    asw_rows: list[tuple[str, Any]] = [
        ("ASW Name", form.asw_name),
        ("ASW Phone", form.asw_phone),
        ("ASW Email", form.asw_email),
    ]

    for i, (label, val) in enumerate(client_rows):
        r = 8 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lc = ws.cell(r, 1, value=label)
        lc.font = _f_sub_header()
        lc.fill = FILL_LIGHT
        lc.alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        vc = ws.cell(r, 3, value=val if val not in (None, "") else "")
        vc.font = _f_input()
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if label.startswith("Pay Rate"):
            vc.number_format = "$#,##0.00"
        _apply_border_range(ws, r, 1, r, 3)
        ws.row_dimensions[r].height = dr

    for i, (label, val) in enumerate(asw_rows):
        r = 8 + i
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        lc = ws.cell(r, 4, value=label)
        lc.font = _f_sub_header()
        lc.fill = FILL_LIGHT
        lc.alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        vc = ws.cell(r, 6, value=val if val not in (None, "") else "")
        vc.font = _f_input()
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        _apply_border_range(ws, r, 4, r, 6)
        ws.row_dimensions[r].height = dr

    for r in (11, 12, 13, 14):
        for c in range(4, 7):
            ws.cell(r, c).value = None

    _spacer_row(ws, 15)

    # --- Authorization (left) vs Scheduled Allocation (right), aligned blocks ---
    subt_r = 16
    _sub_title(ws, "MDHHS AUTHORIZATION", subt_r, 1, 3)
    _sub_title(ws, "SCHEDULED ALLOCATION", subt_r, 4, 6)

    _weekly_min_val = int(sd.get("mdhhs_weekly_minutes") or sd.get("weekly_minutes") or 0)
    _monthly_min_val = int(sd.get("mdhhs_monthly_minutes") or 0)
    auth_rows: list[tuple[str, Any, str]] = [
        ("Weekly Minutes", _weekly_min_val, "0"),
        ("Weekly HH:MM", _fmt_hhmm(_weekly_min_val), "@"),
        ("Monthly Minutes", _monthly_min_val, "0"),
        ("Monthly HH:MM", _fmt_hhmm(_monthly_min_val), "@"),
        (
            "Monthly $",
            float(
                sd.get("mdhhs_form_amount")
                or sd.get("mdhhs_monthly_amount")
                or 0.0
            ),
            "$#,##0.00",
        ),
    ]

    shift_stats = _shift_type_stats(daily, str(catchup_iso) if catchup_iso else None)

    alloc_hdr = 17
    for j, h in enumerate(["Shift Type", "Days", "Total Min"]):
        c = ws.cell(alloc_hdr, 4 + j, value=h)
        c.font = _f_col_header()
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[alloc_hdr].height = float(LAYOUT["header_row_height"])

    # Left card: header row matching allocation column headers so both cards share the same row grid.
    lh = ws.cell(alloc_hdr, 1, value="")
    lh.fill = FILL_HEADER
    lh.font = _f_col_header()
    lh.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=alloc_hdr, start_column=1, end_row=alloc_hdr, end_column=3)
    _apply_border_range(ws, alloc_hdr, 1, alloc_hdr, 3)

    for i, (label, val, fmt) in enumerate(auth_rows):
        r = 18 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lc = ws.cell(r, 1, value=label)
        lc.font = _f_sub_header()
        lc.fill = FILL_LIGHT
        lc.alignment = Alignment(
            horizontal="right", vertical="center", indent=1
        )
        vc = ws.cell(r, 3, value=val)
        vc.font = _f_input()
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.number_format = fmt
        _apply_border_range(ws, r, 1, r, 3)
        ws.row_dimensions[r].height = dr

    alloc_total_min = 0
    for slot in range(5):
        r = 18 + slot
        if slot < len(shift_stats):
            shift_name, n, tot_min, is_catch = shift_stats[slot]
            lc = ws.cell(r, 4, value=shift_name)
            lc.font = _f_body()
            dc = ws.cell(r, 5, value=n)
            dc.font = _f_body()
            dc.alignment = Alignment(horizontal="center")
            mc = ws.cell(r, 6, value=tot_min)
            mc.font = _f_body()
            mc.number_format = "#,##0"
            if is_catch:
                for cc in (4, 5, 6):
                    ws.cell(r, cc).fill = FILL_DEVIATION
            alloc_total_min += tot_min
        _apply_border_range(ws, r, 4, r, 6)
        ws.row_dimensions[r].height = dr

    gr = 23
    lc = ws.cell(gr, 4, value="TOTAL (minutes)")
    lc.font = _f_total()
    lc.fill = FILL_TOTAL
    nc = ws.cell(gr, 5, value=sum(n for _, n, _, _ in shift_stats))
    nc.font = _f_total()
    nc.fill = FILL_TOTAL
    nc.alignment = Alignment(horizontal="center")
    tc = ws.cell(gr, 6, value=alloc_total_min)
    tc.font = _f_total()
    tc.fill = FILL_TOTAL
    tc.number_format = "#,##0"
    _apply_border_range(ws, gr, 4, gr, 6)
    ws.row_dimensions[gr].height = dr

    gr_hm = 24
    lc2 = ws.cell(gr_hm, 4, value="TOTAL HH:MM")
    lc2.font = _f_total()
    lc2.fill = FILL_TOTAL
    ws.cell(gr_hm, 5).fill = FILL_TOTAL
    tc2 = ws.cell(gr_hm, 6, value=_fmt_hhmm(alloc_total_min))
    tc2.font = _f_total()
    tc2.fill = FILL_TOTAL
    tc2.alignment = Alignment(horizontal="right")
    _apply_border_range(ws, gr_hm, 4, gr_hm, 6)
    ws.row_dimensions[gr_hm].height = dr

    # Pad left authorization card to the same vertical span as allocation totals (yellow rows on the right).
    pad_left_total = (gr, gr_hm)
    for pr in pad_left_total:
        corner = ws.cell(pr, 1)
        corner.value = None
        corner.fill = FILL_TOTAL
        corner.font = _f_total()
        ws.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=3)
        _apply_border_range(ws, pr, 1, pr, 3)

    nt_auth, tv_auth = _authorized_non_travel_travel_totals(form)
    exceed_ehhs = nt_auth > EHHS_THRESHOLD_MIN
    ehhs_r1, ehhs_r2 = 25, 26

    lbl1 = ws.cell(ehhs_r1, 1, value="Counts toward 179.59-hr EHHS threshold")
    lbl1.font = _f_sub_header()
    lbl1.fill = FILL_LIGHT
    lbl1.alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    ws.merge_cells(
        start_row=ehhs_r1, start_column=1, end_row=ehhs_r1, end_column=2,
    )
    val1 = (
        f"{nt_auth} min ({_fmt_hours_from_minutes(nt_auth)})"
        + (
            " — EHHS REQUIRED (submit DCH-1785)"
            if exceed_ehhs
            else " — under threshold"
        )
    )
    vc1 = ws.cell(ehhs_r1, 3, value=val1)
    vc1.font = Font(
        name=FONT_NAME,
        size=10,
        color=COLOR_EHHS_EXCEED_FONT if exceed_ehhs else COLOR_EHHS_OK_FONT,
        bold=True,
    )
    vc1.fill = FILL_EHHS_EXCEED if exceed_ehhs else FILL_EHHS_OK
    vc1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(
        start_row=ehhs_r1, start_column=3, end_row=ehhs_r1, end_column=6,
    )
    _apply_border_range(ws, ehhs_r1, 1, ehhs_r1, 6)
    for c in range(1, 7):
        ws.cell(ehhs_r1, c).border = BORDER

    lbl2 = ws.cell(ehhs_r2, 1, value="Excluded (travel for shopping/laundry)")
    lbl2.font = _f_sub_header()
    lbl2.fill = FILL_LIGHT
    lbl2.alignment = Alignment(
        horizontal="right", vertical="center", indent=1
    )
    ws.merge_cells(
        start_row=ehhs_r2, start_column=1, end_row=ehhs_r2, end_column=2,
    )
    vc2 = ws.cell(
        ehhs_r2,
        3,
        value=f"{tv_auth} min ({_fmt_hours_from_minutes(tv_auth)})",
    )
    vc2.font = _f_input()
    vc2.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=ehhs_r2, start_column=3, end_row=ehhs_r2, end_column=6,
    )
    _apply_border_range(ws, ehhs_r2, 1, ehhs_r2, 6)
    for c in range(1, 7):
        ws.cell(ehhs_r2, c).border = BORDER

    _spacer_row(ws, 28)

    recon_row = 29
    _sub_title(ws, "RECONCILIATION", recon_row, 1, 6)

    hdr = recon_row + 1
    h_ab = ws.cell(hdr, 1, value="")
    h_ab.font = _f_sub_header()
    ws.merge_cells(start_row=hdr, start_column=1, end_row=hdr, end_column=2)
    for cc in range(1, 3):
        ws.cell(hdr, cc).fill = FILL_HEADER
    h1 = ws.cell(hdr, 3, value="Minutes")
    h1.font = _f_col_header()
    h1.fill = FILL_HEADER
    h1.alignment = Alignment(horizontal="center", vertical="center")
    h2 = ws.cell(hdr, 4, value="HH:MM")
    h2.font = _f_col_header()
    h2.fill = FILL_HEADER
    h2.alignment = Alignment(horizontal="center", vertical="center")
    h_amt = ws.cell(hdr, 5, value="Amount")
    h_amt.font = _f_col_header()
    h_amt.fill = FILL_HEADER
    h_amt.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=hdr, start_column=5, end_row=hdr, end_column=6)
    ws.cell(hdr, 6).fill = FILL_HEADER
    _apply_border_range(ws, hdr, 1, hdr, 6)
    ws.row_dimensions[hdr].height = float(LAYOUT["header_row_height"])

    recon_rows: list[tuple[str, Any, str, str]] = [
        ("Authorized (cap)", auth_min, _fmt_hhmm(auth_min) or "", f"${auth_amt:,.2f}"),
        ("Delivered (scheduled)", del_min, _fmt_hhmm(del_min) or "", f"${del_amt:,.2f}"),
        ("Billable (cap applied)", bill_min, _fmt_hhmm(bill_min) or "", f"${bill_amt:,.2f}"),
        (tail_label, ov_min, "", _fmt_calendar_overshoot_dollars(ov_amt)),
    ]

    for i, (label, vmin, hhmm, amt_s) in enumerate(recon_rows):
        r = hdr + 1 + i
        lc = ws.cell(r, 1, value=label)
        lc.font = _f_sub_header()
        lc.fill = FILL_LIGHT
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        vc = ws.cell(r, 3, value=vmin)
        vc.number_format = "+0;-0;0" if i == len(recon_rows) - 1 else "#,##0"
        vc.font = _f_xref()
        vc.alignment = Alignment(horizontal="right")
        vh = ws.cell(r, 4, value=hhmm)
        vh.font = _f_xref()
        vh.alignment = Alignment(horizontal="right")
        va = ws.cell(r, 5, value=amt_s)
        va.font = _f_xref()
        va.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        _apply_border_range(ws, r, 1, r, 6)
        ws.row_dimensions[r].height = dr

    stat_r = hdr + 1 + len(recon_rows)
    n_chk = len(vr.checks)
    chk_pass = sum(1 for c in vr.checks if c.passed)
    note_cell = ws.cell(
        stat_r,
        1,
        value=(
            f"Validation checks: {chk_pass}/{n_chk} passed · {tail_note}"
        ),
    )
    note_cell.font = _f_sub_header()
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=stat_r, start_column=1, end_row=stat_r, end_column=4)

    st_cell = ws.cell(stat_r, 5, value=badge)
    st_cell.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_FORMULA)
    st_cell.fill = FILL_TOTAL
    st_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=stat_r, start_column=5, end_row=stat_r, end_column=6)

    _apply_border_range(ws, stat_r, 1, stat_r, 6)
    ws.row_dimensions[stat_r].height = float(LAYOUT["tall_data_row_height"])

    _set_col_widths(ws, {1: 14, 2: 14, 3: 28, 4: 14, 5: 14, 6: 28})
    _set_print_layout(ws, landscape=True, fit_to_width=True, print_title_rows="1:6")
    _professionalize_sheet(ws, title_rows=1, freeze="A7")


def _build_weekly_pattern(ws: Worksheet, sd: dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    month_name = str(sd.get("month_name") or "")
    year = int(sd.get("year") or 0)

    _title_row(
        ws,
        f"Weekly Pattern — {month_name} {year}",
        row=1, col_start=1, col_end=6,
    )

    _col_headers(
        ws, 3,
        ["Day", "Clock In", "Clock Out", "Min (raw)", "HH:MM", "Tasks"],
    )

    wp: dict[str, Any] = dict(sd.get("weekly_pattern") or sd.get("days") or {})
    for i, dow in enumerate(FULL_DOW):
        r = 4 + i
        day_data = wp.get(dow) or {}
        dow_short = SHORT_DOW[i]
        fill = _day_fill(dow_short, is_catchup=False)

        name = ws.cell(r, 1, value=dow)
        name.font = _f_input()
        start = ws.cell(r, 2, value=str(day_data.get("start") or ""))
        start.font = _f_body()
        end = ws.cell(r, 3, value=str(day_data.get("end") or ""))
        end.font = _f_body()
        minutes_val = int(day_data.get("minutes") or 0)
        dur = ws.cell(r, 4, value=minutes_val)
        dur.font = _f_muted()
        dur.number_format = "#,##0"
        dur_hm = ws.cell(r, 5, value=_fmt_hhmm(minutes_val) if minutes_val else "")
        dur_hm.font = _f_body()
        dur_hm.alignment = Alignment(horizontal="right")
        tasks = ", ".join(list(day_data.get("tasks") or []))
        tc = ws.cell(r, 6, value=tasks)
        tc.font = _f_body()
        tc.alignment = Alignment(wrap_text=True, vertical="center")

        for c in range(1, 7):
            if fill is not FILL_NONE:
                ws.cell(r, c).fill = fill
        _apply_border_range(ws, r, 1, r, 6)
        ws.row_dimensions[r].height = float(LAYOUT["tall_data_row_height"])

    total_row = 11
    weekly_total = sum(
        int((wp.get(dow) or {}).get("minutes") or 0) for dow in FULL_DOW
    )
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    total_label = ws.cell(total_row, 1, value="WEEKLY TOTAL")
    total_label.font = _f_total()
    total_label.fill = FILL_TOTAL
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_min = ws.cell(total_row, 4, value=weekly_total)
    total_min.font = _f_total()
    total_min.fill = FILL_TOTAL
    total_min.number_format = "#,##0"
    total_hm = ws.cell(total_row, 5, value=_fmt_hhmm_with_words(weekly_total))
    total_hm.font = _f_total()
    total_hm.fill = FILL_TOTAL
    total_hm.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    total_words = ws.cell(total_row, 6, value=_fmt_hours_minutes_long(weekly_total))
    total_words.font = _f_total()
    total_words.fill = FILL_TOTAL
    total_words.alignment = Alignment(horizontal="left", vertical="center")
    _apply_border_range(ws, total_row, 1, total_row, 6)
    ws.row_dimensions[total_row].height = float(LAYOUT["tall_data_row_height"])

    # Deviation callout below the grid (only if catchup present)
    catchup_iso = sd.get("catchup_date")
    if catchup_iso:
        callout_row = 12
        dt = _parse_iso(str(catchup_iso))
        daily = list(sd.get("daily_schedule") or [])
        catch_entry = next(
            (d for d in daily if d.get("date") == str(catchup_iso)), {}
        )
        # Read the day's actuals straight from daily_schedule so the
        # callout matches exactly what the rest of the workbook shows,
        # regardless of shift shape.
        dur = int(catch_entry.get("duration_min") or 0)
        tasks_text = ", ".join(
            k for k, v in (catch_entry.get("tasks") or {}).items() if v
        )
        clock_in = str(catch_entry.get("clock_in") or "")
        clock_out = str(catch_entry.get("clock_out") or "")
        date_txt = dt.strftime("%A, %B %d, %Y") if dt else str(catchup_iso)
        weekday = dt.strftime("%A") if dt else ""

        _sub_title(ws, "DEVIATION (Catch-up day)", callout_row, 1, 6, fill=FILL_DEVIATION)
        replacement = (
            f"Replaces the normal {weekday} pattern " if weekday else ""
        )
        msg = (
            f"DEVIATION: {date_txt} — {clock_in} to {clock_out}, "
            f"{dur} min ({_fmt_hhmm(dur)}), does {tasks_text}. "
            f"{replacement}to close the calendar-vs-4.3-week gap."
        )
        ws.merge_cells(
            start_row=callout_row + 1, start_column=1,
            end_row=callout_row + 1, end_column=6,
        )
        c = ws.cell(callout_row + 1, 1, value=msg)
        c.font = _f_body()
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = FILL_DEVIATION
        for cc in range(1, 7):
            ws.cell(callout_row + 1, cc).fill = FILL_DEVIATION
            ws.cell(callout_row + 1, cc).border = BORDER
        ws.row_dimensions[callout_row + 1].height = 56

    _set_col_widths(ws, {1: 14, 2: 12, 3: 12, 4: 16, 5: 14, 6: 60})
    _set_print_layout(ws, landscape=True, fit_to_width=True, print_title_rows="1:3")
    _add_excel_table(
        ws,
        ref="A3:F10",
        name=f"{ws.title.replace(' ', '')}Table",
        style=FILL_TABLE_BLUE,
    )
    _professionalize_sheet(ws, title_rows=1, freeze="A4", auto_filter="A3:F11")


# --- Tab 3: Daily Schedule (invariant-checked) -------------------------------


def _build_daily_schedule(
    ws: Worksheet,
    form: ExtractedForm,
    sd: dict[str, Any],
) -> tuple[int, int, int, int]:
    """Return (first_data_row, last_data_row, duration_col, row_total_col)
    so the invariant check can re-read the rendered values afterwards.
    """
    ws.sheet_view.showGridLines = False
    month_name = str(sd.get("month_name") or "")
    year = int(sd.get("year") or 0)

    daily = list(sd.get("daily_schedule") or [])
    task_names = _ordered_task_names(form, daily)
    n_task_cols = len(task_names)

    # Column layout:
    #   1  Date
    #   2  DoW
    #   3  Shift Type
    #   4  Clock In
    #   5  Clock Out
    #   6  Min (raw)             <- Σ task minutes invariant uses this numeric column
    #   7  HH:MM                   <- `_fmt_hhmm(duration)` primary display for reviewers
    #   8..7+n_task_cols  Tasks (minutes, numeric)
    #   8+n_task_cols     Row Total
    dur_hm_col = 7
    last_task_col = 7 + n_task_cols
    row_total_col = last_task_col + 1
    last_col = row_total_col

    _title_row(
        ws,
        f"Daily Schedule — {month_name} {year}",
        row=1, col_start=1, col_end=last_col,
    )

    headers = [
        "Date", "DoW", "Shift Type", "Clock In", "Clock Out",
        "Min (raw)", "HH:MM (+ h/m on week & month totals)",
    ] + task_names + ["Row Total"]
    _col_headers(ws, 3, headers)

    catchup_iso = sd.get("catchup_date")
    day_durations = [int(d.get("duration_min") or 0) for d in daily]
    med_duration = statistics.median(day_durations) if day_durations else 0
    long_thresh = float(med_duration) * 1.5 if med_duration else 0.0

    first_data_row = 4
    current_row = first_data_row
    current_week: int | None = None
    current_week_days: list[dict] = []

    def _emit_weekly_subtotal(at_row: int, days_in_week: list[dict]) -> int:
        if not days_in_week:
            return at_row
        first_date = _parse_iso(str(days_in_week[0].get("date") or ""))
        label = (
            f"Week of {first_date.strftime('%b %d')} — Subtotal"
            if first_date
            else "Week — Subtotal"
        )
        ws.merge_cells(start_row=at_row, start_column=1, end_row=at_row, end_column=5)
        c0 = ws.cell(at_row, 1, value=label)
        c0.font = Font(name=FONT_NAME, size=10, bold=True, italic=True,
                       color=COLOR_FORMULA)
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        wk_total = sum(int(d.get("duration_min") or 0) for d in days_in_week)
        wk_min = ws.cell(at_row, 6, value=wk_total)
        wk_min.font = _f_total()
        wk_min.number_format = "#,##0"

        wk_hm = ws.cell(
            at_row, dur_hm_col, value=_fmt_hhmm_with_words(wk_total)
        )
        wk_hm.font = _f_total()
        wk_hm.alignment = Alignment(horizontal="right", wrap_text=True)

        wk_task_totals: dict[str, int] = {nm: 0 for nm in task_names}
        for d_in in days_in_week:
            for nm, mm in (d_in.get("tasks") or {}).items():
                if nm in wk_task_totals and mm:
                    wk_task_totals[nm] += int(mm)
        for j, nm in enumerate(task_names):
            cell = ws.cell(
                at_row, dur_hm_col + 1 + j,
                value=(wk_task_totals[nm] if wk_task_totals[nm] else None),
            )
            cell.font = _f_total()
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

        wk_rt = ws.cell(at_row, row_total_col, value=wk_total)
        wk_rt.font = _f_total()
        wk_rt.number_format = "#,##0"

        side_med = Side(style="medium", color="2E5C9E")
        side_thin = Side(style="thin", color="C7CFDD")
        for c in range(1, last_col + 1):
            cell = ws.cell(at_row, c)
            cell.fill = FILL_WEEKBAND
            cell.border = Border(
                left=side_thin,
                right=side_thin,
                top=side_med,
                bottom=side_thin,
            )

        ws.row_dimensions[at_row].height = 22
        return at_row + 1

    for d in daily:
        d_date = _parse_iso(str(d.get("date") or ""))
        wk = d_date.isocalendar()[1] if d_date else current_week

        if current_week is not None and wk != current_week:
            current_row = _emit_weekly_subtotal(current_row, current_week_days)
            current_week_days = []

        current_week = wk

        r = current_row
        dow_short = str(d.get("day_of_week") or "")
        is_catchup = bool(catchup_iso) and d.get("date") == str(catchup_iso)
        duration = int(d.get("duration_min") or 0)
        if is_catchup:
            fill: PatternFill = FILL_DEVIATION
        elif day_durations and duration > long_thresh:
            fill = FILL_LIGHT
        else:
            fill = _day_fill(dow_short, is_catchup=False)

        # Core cells
        dt_cell = ws.cell(r, 1, value=str(d.get("date") or ""))
        dt_cell.font = _f_body()
        dow_cell = ws.cell(r, 2, value=dow_short)
        dow_cell.font = _f_body()
        st_cell = ws.cell(r, 3, value=str(d.get("shift_type") or ""))
        st_cell.font = _f_body()

        ci = _ampm_to_time(str(d.get("clock_in") or ""))
        co = _ampm_to_time(str(d.get("clock_out") or ""))
        ci_cell = ws.cell(r, 4, value=ci if isinstance(ci, time) else str(ci))
        if isinstance(ci, time):
            ci_cell.number_format = "h:mm AM/PM;@"
        ci_cell.font = _f_body()
        co_cell = ws.cell(r, 5, value=co if isinstance(co, time) else str(co))
        if isinstance(co, time):
            co_cell.number_format = "h:mm AM/PM;@"
        co_cell.font = _f_body()

        dur_cell = ws.cell(r, 6, value=duration)
        dur_cell.font = _f_muted()
        dur_cell.number_format = "#,##0"
        dur_hm_cell = ws.cell(r, dur_hm_col, value=_fmt_hhmm(duration))
        dur_hm_cell.font = _f_body()
        dur_hm_cell.alignment = Alignment(horizontal="right")

        # Task minute cells: literal ints from DayEntry.tasks (source of truth).
        task_map: dict[str, int] = dict(d.get("tasks") or {})
        row_sum = 0
        for j, nm in enumerate(task_names):
            c = dur_hm_col + 1 + j
            m = int(task_map.get(nm, 0) or 0)
            cell = ws.cell(r, c, value=(m if m > 0 else None))
            cell.font = _f_body()
            cell.alignment = Alignment(horizontal="right")
            if m > 0:
                row_sum += m

        # Row total: literal python sum (so data_only readers don't need recalc;
        # openpyxl would otherwise write formula without cached value).
        rt_cell = ws.cell(r, row_total_col, value=row_sum)
        rt_cell.font = _f_total()
        rt_cell.number_format = "#,##0"
        rt_cell.fill = FILL_TOTAL

        # Apply day fill to all columns except the row-total (which keeps the
        # yellow totals color).
        if fill is not FILL_NONE:
            for c in range(1, row_total_col):
                ws.cell(r, c).fill = fill

        _apply_border_range(ws, r, 1, r, last_col)

        current_week_days.append(d)
        current_row += 1

    if current_week_days:
        current_row = _emit_weekly_subtotal(current_row, current_week_days)

    last_data_row = current_row - 1

    # --- TOTALS row (literal sums from the daily schedule) ---
    tot_row = last_data_row + 1
    for c in range(1, last_col + 1):
        ws.cell(tot_row, c).fill = FILL_TOTAL
        ws.cell(tot_row, c).border = BORDER

    label = ws.cell(tot_row, 1, value="MONTHLY TOTAL")
    label.font = _f_total()
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)

    total_duration = sum(int(d.get("duration_min") or 0) for d in daily)
    traw = ws.cell(tot_row, 6, value=total_duration)
    traw.font = _f_total()
    traw.number_format = "#,##0"
    tot_hm_cell = ws.cell(
        tot_row, dur_hm_col, value=_fmt_hhmm_with_words(total_duration)
    )
    tot_hm_cell.font = _f_total()
    tot_hm_cell.alignment = Alignment(horizontal="right", wrap_text=True)

    task_totals: dict[str, int] = {nm: 0 for nm in task_names}
    for d in daily:
        for nm, m in (d.get("tasks") or {}).items():
            if nm in task_totals and m:
                task_totals[nm] += int(m)
    for j, nm in enumerate(task_names):
        cell = ws.cell(tot_row, dur_hm_col + 1 + j, value=task_totals[nm])
        cell.font = _f_total()
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")

    grand = sum(task_totals.values())
    gt = ws.cell(tot_row, row_total_col, value=grand)
    gt.font = _f_total()
    gt.number_format = "#,##0"

    # --- Per-task HH:MM (display-only mirror; invariant uses numeric grid above) ---
    pm_title = tot_row + 2
    _sub_title(ws, "Per-Task HH:MM (display only)", pm_title, 1, last_col)
    pm_hdr = pm_title + 1
    ph0: list[str] = (
        ["Date", "—", "—", "—", "—", "—", "—"]
        + [f"{nm}" for nm in task_names]
        + ["Row HH:MM"]
    )
    _col_headers(ws, pm_hdr, ph0)
    pm_first = pm_hdr + 1
    for i, d in enumerate(daily):
        r = pm_first + i
        task_map = dict(d.get("tasks") or {})
        row_sum = 0
        ws.cell(r, 1, value=str(d.get("date") or "")).font = _f_body()
        for cx in range(2, dur_hm_col + 1):
            ws.cell(r, cx, value="").font = _f_body()
        for j, nm in enumerate(task_names):
            m = int(task_map.get(nm, 0) or 0)
            row_sum += m
            c = ws.cell(
                r,
                dur_hm_col + 1 + j,
                value=_fmt_hhmm(m) if m > 0 else "",
            )
            c.font = _f_body()
            c.alignment = Alignment(horizontal="right")
        ws.cell(
            r, row_total_col, value=_fmt_hhmm(row_sum) if row_sum else ""
        ).font = _f_body()
        _apply_border_range(ws, r, 1, r, last_col)

    # --- Column widths (uniform task minute columns) ---
    tw = int(LAYOUT["task_min_col_width"])
    widths = {
        1: 12,
        2: 12,
        3: 14,
        4: 11,
        5: 11,
        6: 10,
        dur_hm_col: 16,
    }
    for j in range(n_task_cols):
        widths[dur_hm_col + 1 + j] = float(tw)
    widths[row_total_col] = 12
    _set_col_widths(ws, widths)
    ws.freeze_panes = "A4"
    _set_print_layout(ws, landscape=False, fit_to_width=True, print_title_rows="1:3")
    _professionalize_sheet(
        ws,
        title_rows=1,
        freeze="A4",
        auto_filter=f"A3:{get_column_letter(last_col)}{last_data_row}",
    )

    return first_data_row, last_data_row, 6, row_total_col


def _assert_daily_row_invariant(
    ws: Worksheet,
    first_row: int,
    last_row: int,
    duration_col: int,
    row_total_col: int,
) -> None:
    """Σ(task minute cells in a row) MUST equal the duration cell.

    Task minute cells live in columns [duration_col+1 .. row_total_col-1].
    The row-total cell must match both sums.
    """
    for r in range(first_row, last_row + 1):
        duration = ws.cell(r, duration_col).value or 0
        row_total = ws.cell(r, row_total_col).value or 0
        task_sum = 0
        for c in range(duration_col + 1, row_total_col):
            v = ws.cell(r, c).value
            # Skip non-numeric companion cells (e.g. ``HH:MM`` text column) — those are display-only and not part of the
            # task-minute payload.
            if v is None or not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            task_sum += int(v)
        if int(task_sum) != int(duration) or int(row_total) != int(duration):
            raise AssertionError(
                f"Daily Schedule row {r}: Σ(tasks)={task_sum}, "
                f"row_total={row_total}, duration={duration} — mismatch. "
                "The xlsx renderer would drift from the month-aware schedule."
            )


# --- Tab 4: Schedule Math ----------------------------------------------------


def _build_schedule_math(ws: Worksheet, sd: dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    month_name = str(sd.get("month_name") or "")
    year = int(sd.get("year") or 0)

    _title_row(
        ws,
        f"Schedule Math — {month_name} {year}",
        row=1, col_start=1, col_end=6,
    )

    # --- MDHHS 4.3-week derivation ---
    weekly = int(
        sd.get("mdhhs_weekly_minutes") or sd.get("weekly_minutes") or 0
    )
    monthly_target = int(sd.get("mdhhs_monthly_minutes") or 0)
    monthly_amount = float(sd.get("mdhhs_monthly_amount") or 0.0)

    _sub_title(ws, "MDHHS 4.3-WEEK DERIVATION", 3, 1, 6)
    rows: list[tuple[str, Any, str]] = [
        ("Weekly minutes (Σ min/day × days/wk)", weekly, "#,##0"),
        ("Weekly HH:MM", _fmt_hhmm(weekly), "@"),
        ("Weekly Hours + Minutes", _fmt_hours_minutes_long(weekly), "@"),
        ("Weeks per month (MDHHS rule)", WEEKS_PER_MONTH, "0.0"),
        ("Monthly target (weekly × 4.3, rounded)", monthly_target, "#,##0"),
        ("Monthly HH:MM", _fmt_hhmm(monthly_target), "@"),
        ("Monthly Hours + Minutes", _fmt_hours_minutes_long(monthly_target), "@"),
        ("Monthly $ target", monthly_amount, "$#,##0.00"),
    ]
    for i, (label, val, fmt) in enumerate(rows):
        r = 4 + i
        lc = ws.cell(r, 1, value=label)
        lc.font = _f_sub_header()
        lc.fill = FILL_LIGHT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        vc = ws.cell(r, 4, value=val)
        vc.font = _f_input()
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        _apply_border_range(ws, r, 1, r, 6)
        ws.row_dimensions[r].height = float(LAYOUT["data_row_height"])

    _spacer_row(ws, 10)

    # --- Shift-type breakdown ---
    # Aggregated from ``daily_schedule`` — no hardcoded shift-shape
    # assumptions about which weekdays host which frequencies. Every
    # shift_type the scheduler emitted gets one row showing days and
    # total minutes, ordered by first appearance.
    daily = list(sd.get("daily_schedule") or [])
    catchup_iso = sd.get("catchup_date")
    shift_stats = _shift_type_stats(
        daily, str(catchup_iso) if catchup_iso else None
    )

    break_row = 4 + len(rows) + 1  # 1-row gap below the derivation table
    _sub_title(ws, "SHIFT-TYPE BREAKDOWN", break_row, 1, 6)
    _col_headers(
        ws, break_row + 1,
        ["Shift Type", "Days", "Total Minutes", "Total HH:MM", "Avg Min / Day", "Running"],
    )

    running = 0
    for i, (shift_name, n, tot_min, is_catch) in enumerate(shift_stats):
        r = break_row + 2 + i
        running += tot_min
        fill = FILL_DEVIATION if is_catch else FILL_NONE
        avg = round_half_up(tot_min / n) if n > 0 else 0
        cells: list[tuple[int, Any, Any]] = [
            (1, shift_name, _f_body()),
            (2, n, _f_body()),
            (3, tot_min, _f_body()),
            (4, _fmt_hhmm(tot_min), _f_body()),
            (5, avg, _f_body()),
            (6, running, _f_body()),
        ]
        for col, val, fnt in cells:
            c = ws.cell(r, col, value=val)
            c.font = fnt
            if isinstance(val, int) and col not in (1, 4):
                c.number_format = "#,##0"
            if col == 2:
                c.alignment = Alignment(horizontal="center")
            if col == 4:
                c.alignment = Alignment(horizontal="right")
            if col in (3, 5, 6):
                c.alignment = Alignment(horizontal="right")
            if fill is not FILL_NONE:
                c.fill = fill
        _apply_border_range(ws, r, 1, r, 6)
        ws.row_dimensions[r].height = float(LAYOUT["data_row_height"])
    gt_row = break_row + 2 + max(len(shift_stats), 1)
    for c in range(1, 7):
        ws.cell(gt_row, c).fill = FILL_TOTAL
        ws.cell(gt_row, c).border = BORDER
    lbl = ws.cell(gt_row, 1, value="TOTAL (monthly minutes)")
    lbl.font = _f_total()
    tot_days = ws.cell(gt_row, 2, value=sum(n for _, n, _, _ in shift_stats))
    tot_days.font = _f_total()
    tot_days.alignment = Alignment(horizontal="center")
    tot_min_cell = ws.cell(gt_row, 3, value=running)
    tot_min_cell.font = _f_total()
    tot_min_cell.number_format = "#,##0"
    tot_hm_cell = ws.cell(gt_row, 4, value=_fmt_hhmm(running))
    tot_hm_cell.font = _f_total()
    tot_hm_cell.alignment = Alignment(horizontal="right")
    tot_run = ws.cell(gt_row, 6, value=running)
    tot_run.font = _f_total()
    tot_run.number_format = "#,##0"
    tot_run.alignment = Alignment(horizontal="right")

    _spacer_row(ws, gt_row + 1)

    # --- 4.3-week cumulative reconciliation ---
    cum = list(sd.get("cumulative_by_day") or [])
    if cum:
        cum_row = gt_row + 2
        form_amt = float(
            sd.get("mdhhs_form_amount")
            or sd.get("mdhhs_monthly_amount")
            or monthly_amount
        )
        cap_s = f"${form_amt:,.2f}"
        _sub_title(
            ws,
            f"DAILY CUMULATIVE (calendar projection — billable capped at {cap_s})",
            cum_row,
            1,
            5,
        )
        _col_headers(
            ws, cum_row + 1,
            [
                "Date",
                "Daily Minutes",
                "Daily HH:MM",
                "Cumulative",
                "Cumulative HH:MM",
            ],
        )
        for i, entry in enumerate(cum):
            if len(entry) < 4:
                continue
            r = cum_row + 2 + i
            dt_s = entry[0]
            daily_m = int(entry[1])
            cum_m = int(entry[2])
            is_catch = (
                sd.get("catchup_date") is not None and dt_s == str(sd.get("catchup_date"))
            )
            if is_catch:
                fill = FILL_DEVIATION
            elif (i + 1) % 2 == 0:
                fill = FILL_LIGHT
            else:
                fill = FILL_NONE

            c1 = ws.cell(r, 1, value=str(dt_s))
            c1.font = _f_body()
            c2 = ws.cell(r, 2, value=daily_m)
            c2.font = _f_body()
            c2.number_format = "#,##0"
            c2.alignment = Alignment(horizontal="right")
            dm_hm = ws.cell(r, 3, value=_fmt_hhmm(daily_m))
            dm_hm.font = _f_body()
            dm_hm.alignment = Alignment(horizontal="right")
            c4 = ws.cell(r, 4, value=cum_m)
            c4.font = _f_body()
            c4.number_format = "#,##0"
            c4.alignment = Alignment(horizontal="right")
            cum_hm = ws.cell(r, 5, value=_fmt_hhmm(cum_m))
            cum_hm.font = _f_body()
            cum_hm.alignment = Alignment(horizontal="right")
            if fill is not FILL_NONE:
                for c in range(1, 6):
                    ws.cell(r, c).fill = fill
            _apply_border_range(ws, r, 1, r, 5)
            ws.row_dimensions[r].height = float(LAYOUT["data_row_height"])

        ws.freeze_panes = f"A{cum_row + 2}"
        bottom_row = cum_row + 2 + len(cum)
        for c in range(1, 6):
            ws.cell(bottom_row, c).fill = FILL_TOTAL
            ws.cell(bottom_row, c).border = BORDER
        ws.merge_cells(
            start_row=bottom_row, start_column=1,
            end_row=bottom_row, end_column=3,
        )
        lbl = ws.cell(bottom_row, 1, value="MONTH CLOSES AT")
        lbl.font = _f_total()
        vc = ws.cell(bottom_row, 4, value=running)
        vc.font = _f_total()
        vc.number_format = "#,##0"
        vc.alignment = Alignment(horizontal="right")
        tcell = ws.cell(bottom_row, 5, value=_fmt_hhmm(running))
        tcell.font = _f_total()
        tcell.alignment = Alignment(horizontal="right")

        # --- Line chart: cumulative delivered vs authorized monthly cap (helper col T) ---
        missing_chart = (
            LineChart,
            Reference,
            GraphicalProperties,
            LineProperties,
            SeriesLabel,
        )
        if any(x is None for x in missing_chart):
            logger.warning(
                "Schedule Math daily cumulative chart skipped: openpyxl chart components "
                "not available"
            )
        else:
            try:
                cap_col = 20
                auth_min_total = int(sd.get("mdhhs_monthly_minutes") or 0)
                cum_first = cum_row + 2
                days_in_month = len(daily)

                if days_in_month > 0:
                    ws.cell(cum_first - 1, cap_col, value="Auth Cap")
                    for i in range(days_in_month):
                        ws.cell(cum_first + i, cap_col, value=auth_min_total)
                    ws.column_dimensions[get_column_letter(cap_col)].hidden = True

                    chart = LineChart()
                    chart.title = "Daily Cumulative Delivery vs Authorized Cap"
                    chart.style = 2
                    chart.y_axis.title = "Cumulative Minutes"
                    chart.x_axis.title = "Day"
                    chart.height = 9
                    chart.width = 22

                    cum_min_col = 4
                    cum_ref = Reference(
                        ws,
                        min_col=cum_min_col,
                        min_row=cum_first,
                        max_col=cum_min_col,
                        max_row=cum_first + days_in_month - 1,
                    )
                    chart.add_data(cum_ref, titles_from_data=False)
                    chart.series[0].tx = SeriesLabel(v="Cumulative Delivered")

                    cap_ref = Reference(
                        ws,
                        min_col=cap_col,
                        min_row=cum_first - 1,
                        max_col=cap_col,
                        max_row=cum_first + days_in_month - 1,
                    )
                    chart.add_data(cap_ref, titles_from_data=True)

                    cats = Reference(
                        ws,
                        min_col=1,
                        min_row=cum_first,
                        max_col=1,
                        max_row=cum_first + days_in_month - 1,
                    )
                    chart.set_categories(cats)

                    chart.series[0].graphicalProperties = GraphicalProperties(
                        ln=LineProperties(solidFill="2E5C9E", w=28000),
                    )
                    chart.series[1].graphicalProperties = GraphicalProperties(
                        ln=LineProperties(
                            solidFill="C0392B",
                            w=18000,
                            prstDash="dash",
                        ),
                    )
                    chart.legend.position = "b"

                    ws.add_chart(chart, "H4")
            except Exception as e:
                logger.warning("Schedule Math daily cumulative chart skipped: %s", e)

    _set_col_widths(ws, {1: 38, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14})
    _set_print_layout(ws, landscape=True, fit_to_width=True, print_title_rows="1:3")
    _professionalize_sheet(ws, title_rows=1)


def _write_task_reconciliation_subtotal_row(
    ws: Worksheet,
    row: int,
    last_col: int,
    *,
    label: str,
    row_fill: PatternFill,
    auth_min: int,
    auth_amt: float,
    occ_sum: int,
    sched_min: int,
    variance: int,
    sched_total: int,
    auth_min_col: int,
    auth_hm_col: int,
    auth_amt_col: int,
    occ_col: int,
    sched_min_col: int,
    sched_hm_col: int,
    var_col: int,
    variance_hm_col: int,
    pct_share_col: int,
) -> None:
    """Muted totals row (italic, lighter fill) for EHHS vs travel splits."""
    fnt_num = _f_total_muted(italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c0 = ws.cell(row, 1, value=label)
    c0.font = _f_total_muted(italic=True)
    c0.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c0.fill = row_fill

    ws.cell(row, auth_min_col, value=auth_min).font = fnt_num
    ws.cell(row, auth_min_col).number_format = "#,##0"
    ws.cell(row, auth_min_col).fill = row_fill
    auth_hm = ws.cell(
        row, auth_hm_col, value=_fmt_hhmm(auth_min) if auth_min else ""
    )
    auth_hm.font = fnt_num
    auth_hm.alignment = Alignment(horizontal="right")
    auth_hm.fill = row_fill

    amt = ws.cell(row, auth_amt_col, value=auth_amt)
    amt.font = fnt_num
    amt.number_format = '"$"#,##0.00'
    amt.alignment = Alignment(horizontal="right")
    amt.fill = row_fill

    ws.cell(row, occ_col, value=occ_sum).font = fnt_num
    ws.cell(row, occ_col).alignment = Alignment(horizontal="center")
    ws.cell(row, occ_col).fill = row_fill

    ws.cell(row, sched_min_col, value=sched_min).font = fnt_num
    ws.cell(row, sched_min_col).number_format = "#,##0"
    ws.cell(row, sched_min_col).fill = row_fill
    sched_hm = ws.cell(
        row, sched_hm_col, value=_fmt_hhmm(sched_min) if sched_min else ""
    )
    sched_hm.font = fnt_num
    sched_hm.alignment = Alignment(horizontal="right")
    sched_hm.fill = row_fill

    ws.cell(row, var_col, value=variance).font = fnt_num
    ws.cell(row, var_col).number_format = "+#,##0;-#,##0;0"
    ws.cell(row, var_col).fill = row_fill

    wm = ws.cell(
        row, variance_hm_col, value=_fmt_hhmm(variance) if variance else ""
    )
    wm.font = fnt_num
    wm.fill = row_fill
    wm.alignment = Alignment(horizontal="right")

    pct = sched_min / max(1, sched_total) if sched_total else 0.0
    pcell = ws.cell(row, pct_share_col, value=pct)
    pcell.font = fnt_num
    pcell.number_format = "0.0%"
    pcell.alignment = Alignment(horizontal="right")
    pcell.fill = row_fill

    for c in range(1, last_col + 1):
        ws.cell(row, c).border = BORDER


# --- Tab 5: Task Reconciliation ---------------------------------------------

def _build_task_reconciliation(
    ws: Worksheet,
    form: ExtractedForm,
    sd: dict[str, Any],
    vr: ValidationReport,
) -> None:
    """Render the per-task reconciliation directly from the schedule dict.

    Data sources — this tab does not re-derive anything from frequency
    tables:

    * **Scheduled minutes** per task come from summing
      ``sd["daily_schedule"][*].tasks[name]``.
    * **Occurrence counts** per task come from
      ``sd["task_occurrence_counts"]``, with a live recount from
      ``daily_schedule`` as a fallback when that field is missing
      (legacy payloads).

    Authorised minutes use the standard MDHHS 4.3-week math on the
    extracted form's ``min_per_day × days_per_week``. The grand-total
    variance reflects calendar shape (month length) vs the 4.3-week
    authorization target; per ASM 144, billable amounts still cap at
    authorized dollars.
    """
    ws.sheet_view.showGridLines = False
    month_name = str(sd.get("month_name") or "")
    year = int(sd.get("year") or 0)
    # Column layout (13 cols):
    #   1 Task | 2 Min/Day | 3 Day HH:MM | 4 Days/Wk
    #   5 Auth Min (× 4.3) | 6 Auth HH:MM | 7 Auth $
    #   8 Occurrences | 9 Scheduled Min | 10 Sched HH:MM
    #   11 Variance | 12 Variance HH:MM | 13 % Share
    last_col = 13
    DPW_COL = 4
    AUTH_MIN_COL = 5
    AUTH_HM_COL = 6
    AUTH_AMT_COL = 7
    OCC_COL = 8
    SCHED_MIN_COL = 9
    SCHED_HM_COL = 10
    VAR_COL = 11
    VAR_HM_COL = 12
    PCT_SHARE_COL = 13

    _title_row(
        ws,
        f"Task Reconciliation — {month_name} {year}",
        row=1, col_start=1, col_end=last_col,
    )

    _col_headers(
        ws, 3,
        [
            "Task",
            "Min/Day",
            "Day HH:MM",
            "Days/Wk",
            "Auth Min (× 4.3)",
            "Auth HH:MM",
            "Auth $",
            "Occurrences",
            "Scheduled Min",
            "Sched HH:MM",
            "Variance",
            "Variance HH:MM",
            "% Share",
        ],
    )

    # Aggregate scheduled minutes per task from the source-of-truth daily list.
    daily = list(sd.get("daily_schedule") or [])
    sched_by_task: dict[str, int] = {}
    rebuilt_counts: dict[str, int] = {}
    for d in daily:
        for nm, m in (d.get("tasks") or {}).items():
            mi = int(m or 0)
            if mi > 0:
                sched_by_task[nm] = sched_by_task.get(nm, 0) + mi
                rebuilt_counts[nm] = rebuilt_counts.get(nm, 0) + 1

    # Prefer the schedule dict's declared occurrence map; fall back to
    # a live recount from daily_schedule when the field is absent
    # (older payloads predating task_occurrence_counts).
    declared_counts = sd.get("task_occurrence_counts")
    if isinstance(declared_counts, dict):
        occ_counts: dict[str, int] = {
            str(k): int(v) for k, v in declared_counts.items() if isinstance(v, (int, float))
        }
    else:
        occ_counts = rebuilt_counts

    # Preserve extracted-form task order and add any stragglers from schedule.
    ordered = _ordered_task_names(form, daily)
    task_meta: dict[str, tuple[int, int]] = {}
    for t in form.tasks:
        nm = str(t.get("task_name") or "")
        if nm:
            task_meta[nm] = (int(t.get("min_per_day") or 0), int(t.get("days_per_week") or 0))

    first_data_row = 4
    pay_sr = float(sd.get("pay_rate") or getattr(form, "pay_rate", 0.0) or 0.0)
    sched_total = 0
    se_auth = se_occ = se_sched = se_var = 0
    se_auth_amt = 0.0
    st_auth = st_occ = st_sched = st_var = 0
    st_auth_amt = 0.0
    travel_excl = TRAVEL_TASKS_EXCLUDED_FROM_EHHS

    row_payloads: list[dict[str, Any]] = []
    for nm in ordered:
        mpd, dpw = task_meta.get(nm, (0, 0))
        auth_min = compute_monthly_minutes_rounded(mpd, dpw)
        sched_min = int(sched_by_task.get(nm, 0))
        occ = int(occ_counts.get(nm, 0))
        var = sched_min - auth_min
        auth_amt = (
            compute_task_amount(mpd, dpw, pay_sr)
            if (mpd and dpw and pay_sr)
            else 0.0
        )
        row_payloads.append(
            {
                "nm": nm,
                "mpd": mpd,
                "dpw": dpw,
                "auth_min": auth_min,
                "sched_min": sched_min,
                "occ": occ,
                "var": var,
                "auth_amt": auth_amt,
            }
        )
        sched_total += sched_min
        if nm in travel_excl:
            st_auth += auth_min
            st_occ += occ
            st_sched += sched_min
            st_var += var
            st_auth_amt += auth_amt
        else:
            se_auth += auth_min
            se_occ += occ
            se_sched += sched_min
            se_var += var
            se_auth_amt += auth_amt

    for i, rp in enumerate(row_payloads):
        r = first_data_row + i
        nm = str(rp["nm"])
        mpd = int(rp["mpd"])
        dpw = int(rp["dpw"])
        auth_min = int(rp["auth_min"])
        sched_min = int(rp["sched_min"])
        occ = int(rp["occ"])
        var = int(rp["var"])
        auth_amt = float(rp["auth_amt"])
        share = sched_min / max(1, sched_total) if sched_total else 0.0

        ws.cell(r, 1, value=nm).font = _f_input()
        ws.cell(r, 2, value=mpd).font = _f_body()
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        day_h = ws.cell(r, 3, value=_fmt_hhmm(mpd) if mpd else "")
        day_h.font = _f_body()
        day_h.alignment = Alignment(horizontal="right")
        ws.cell(r, DPW_COL, value=dpw).font = _f_body()
        ws.cell(r, DPW_COL).alignment = Alignment(horizontal="center")
        ws.cell(r, AUTH_MIN_COL, value=auth_min).font = _f_body()
        ws.cell(r, AUTH_MIN_COL).number_format = "#,##0"
        auth_hm = ws.cell(
            r, AUTH_HM_COL, value=_fmt_hhmm(auth_min) if auth_min else ""
        )
        auth_hm.font = _f_body()
        auth_hm.alignment = Alignment(horizontal="right")

        amt_cell = ws.cell(r, AUTH_AMT_COL, value=auth_amt)
        amt_cell.font = _f_body()
        amt_cell.number_format = '"$"#,##0.00'
        amt_cell.alignment = Alignment(horizontal="right")

        ws.cell(r, OCC_COL, value=occ).font = _f_body()
        ws.cell(r, OCC_COL).alignment = Alignment(horizontal="center")
        ws.cell(r, SCHED_MIN_COL, value=sched_min).font = _f_body()
        ws.cell(r, SCHED_MIN_COL).number_format = "#,##0"
        sched_hm = ws.cell(
            r, SCHED_HM_COL, value=_fmt_hhmm(sched_min) if sched_min else ""
        )
        sched_hm.font = _f_body()
        sched_hm.alignment = Alignment(horizontal="right")
        vcell = ws.cell(r, VAR_COL, value=var)
        vcell.font = _f_body()
        vcell.number_format = "+#,##0;-#,##0;0"
        vcell.fill = FILL_VARIANCE_OK if var >= 0 else FILL_VARIANCE_NEG
        vhm = ws.cell(r, VAR_HM_COL, value=_fmt_hhmm(var) if var else "")
        vhm.font = _f_body()
        vhm.alignment = Alignment(horizontal="right")

        pct_cell = ws.cell(r, PCT_SHARE_COL, value=share)
        pct_cell.font = _f_muted()
        pct_cell.number_format = "0.0%"
        pct_cell.alignment = Alignment(horizontal="right")

        _apply_border_range(ws, r, 1, r, last_col)

    last_data_row = first_data_row + len(ordered) - 1

    # --- Subtotals (EHHS-bound vs travel; travel excluded from 179.59-hr cap) ---
    sub_ehhs_row = last_data_row + 1
    sub_tr_row = last_data_row + 2
    _write_task_reconciliation_subtotal_row(
        ws,
        sub_ehhs_row,
        last_col,
        label=(
            "Subtotal — Personal Care + IADL + Complex Care (counts toward EHHS)"
        ),
        row_fill=FILL_SUBTOTAL_EHHS,
        auth_min=se_auth,
        auth_amt=se_auth_amt,
        occ_sum=se_occ,
        sched_min=se_sched,
        variance=se_var,
        sched_total=sched_total,
        auth_min_col=AUTH_MIN_COL,
        auth_hm_col=AUTH_HM_COL,
        auth_amt_col=AUTH_AMT_COL,
        occ_col=OCC_COL,
        sched_min_col=SCHED_MIN_COL,
        sched_hm_col=SCHED_HM_COL,
        var_col=VAR_COL,
        variance_hm_col=VAR_HM_COL,
        pct_share_col=PCT_SHARE_COL,
    )
    _write_task_reconciliation_subtotal_row(
        ws,
        sub_tr_row,
        last_col,
        label="Subtotal — Travel (excluded from 179.59-hr EHHS cap)",
        row_fill=FILL_SUBTOTAL_TRAVEL,
        auth_min=st_auth,
        auth_amt=st_auth_amt,
        occ_sum=st_occ,
        sched_min=st_sched,
        variance=st_var,
        sched_total=sched_total,
        auth_min_col=AUTH_MIN_COL,
        auth_hm_col=AUTH_HM_COL,
        auth_amt_col=AUTH_AMT_COL,
        occ_col=OCC_COL,
        sched_min_col=SCHED_MIN_COL,
        sched_hm_col=SCHED_HM_COL,
        var_col=VAR_COL,
        variance_hm_col=VAR_HM_COL,
        pct_share_col=PCT_SHARE_COL,
    )

    # --- Grand-total row: variance is informational; the status badge
    # below reflects the validator's verdict (which carries the one-
    # session tolerance from validate.py Check 3), not strict equality. ---
    target_monthly_min = int(sd.get("mdhhs_monthly_minutes") or 0)
    total_occurrences = sum(int(v) for v in occ_counts.values())
    auth_total_amt = compute_mdhhs_form_amount(
        [
            {"min_per_day": int(t.get("min_per_day") or 0), "days_per_week": int(t.get("days_per_week") or 0)}
            for t in form.tasks
        ],
        pay_sr,
    )

    gt_row = sub_tr_row + 1
    for c in range(1, last_col + 1):
        ws.cell(gt_row, c).fill = FILL_TOTAL
        ws.cell(gt_row, c).border = BORDER
    ws.cell(gt_row, 1, value="TOTAL").font = _f_total()
    tot_auth_cell = ws.cell(gt_row, AUTH_MIN_COL, value=target_monthly_min)
    tot_auth_cell.font = _f_total()
    tot_auth_cell.number_format = "#,##0"
    tot_auth_hm = ws.cell(gt_row, AUTH_HM_COL, value=_fmt_hhmm(target_monthly_min))
    tot_auth_hm.font = _f_total()
    tot_auth_hm.alignment = Alignment(horizontal="right")

    tot_auth_amt_cell = ws.cell(gt_row, AUTH_AMT_COL, value=auth_total_amt)
    tot_auth_amt_cell.font = _f_total()
    tot_auth_amt_cell.number_format = '"$"#,##0.00'
    tot_auth_amt_cell.alignment = Alignment(horizontal="right")

    tot_occ_cell = ws.cell(gt_row, OCC_COL, value=total_occurrences)
    tot_occ_cell.font = _f_total()
    tot_occ_cell.alignment = Alignment(horizontal="center")
    tot_sched_cell = ws.cell(gt_row, SCHED_MIN_COL, value=sched_total)
    tot_sched_cell.font = _f_total()
    tot_sched_cell.number_format = "#,##0"
    tot_sched_hm = ws.cell(gt_row, SCHED_HM_COL, value=_fmt_hhmm(sched_total))
    tot_sched_hm.font = _f_total()
    tot_sched_hm.alignment = Alignment(horizontal="right")
    grand_var = sched_total - target_monthly_min
    var_cell = ws.cell(gt_row, VAR_COL, value=grand_var)
    var_cell.font = _f_total()
    var_cell.number_format = "+#,##0;-#,##0;0"
    var_hm_cell = ws.cell(gt_row, VAR_HM_COL, value=_fmt_hhmm(grand_var))
    var_hm_cell.font = _f_total()
    var_hm_cell.alignment = Alignment(horizontal="right")

    pct_total = ws.cell(gt_row, PCT_SHARE_COL, value=1.0)
    pct_total.font = _f_total()
    pct_total.number_format = "0.0%"
    pct_total.alignment = Alignment(horizontal="right")
    _add_excel_table(
        ws,
        ref=f"A3:{get_column_letter(last_col)}{gt_row}",
        name="TaskReconciliationTable",
        style=FILL_TABLE_GREEN,
    )

    # --- Status row ---
    status_row = gt_row + 1
    st_labels = {
        "BILLABLE_EXACT": "BILLABLE EXACT ✓",
        "BILLABLE_AT_CAP": "BILLABLE AT CAP ✓",
        "BILLABLE_UNDER_CAP": "BILLABLE UNDER CAP ✓",
        "INVALID": "INVALID ✗",
    }
    vstat = getattr(vr, "validation_status", "INVALID")
    status_text = st_labels.get(vstat, vstat)
    ws.merge_cells(
        start_row=status_row, start_column=1,
        end_row=status_row, end_column=last_col - 1,
    )
    lbl = ws.cell(status_row, 1, value="Status")
    lbl.font = _f_total()
    lbl.fill = FILL_TOTAL
    lbl.alignment = Alignment(horizontal="right")
    sc = ws.cell(status_row, last_col, value=status_text)
    sc.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_FORMULA)
    sc.fill = FILL_TOTAL
    sc.alignment = Alignment(horizontal="center")
    for c in range(1, last_col + 1):
        ws.cell(status_row, c).border = BORDER

    # --- Validation checks summary below ---
    chk_row = status_row + 2
    _sub_title(ws, f"VALIDATION CHECKS ({len(vr.checks)})", chk_row, 1, last_col)
    _col_headers(
        ws, chk_row + 1,
        ["#", "Check", "Status", "Expected", "Actual", "Detail"],
    )
    for i, ch in enumerate(vr.checks):
        r = chk_row + 2 + i
        ws.cell(r, 1, value=int(ch.number)).font = _f_body()
        ws.cell(r, 2, value=str(ch.name)).font = _f_body()
        stat = ws.cell(r, 3, value=("PASS" if ch.passed else "FAIL"))
        stat.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_FORMULA)
        stat.fill = FILL_TOTAL if ch.passed else FILL_DEVIATION
        stat.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, value=_short(str(ch.expected), 120)).font = _f_body()
        ws.cell(r, 5, value=_short(str(ch.actual), 120)).font = _f_body()
        dc = ws.cell(r, 6, value=str(ch.detail or "").replace("\n", " "))
        dc.font = _f_body()
        dc.alignment = Alignment(wrap_text=True, vertical="top")
        _apply_border_range(ws, r, 1, r, last_col)
        ws.row_dimensions[r].height = float(LAYOUT["tall_data_row_height"])

    _set_col_widths(
        ws,
        {
            1: 6,
            2: 24,
            3: 12,
            4: 12,
            5: 18,
            6: 60,
            7: 11,
            8: 14,
            9: 18,
            10: 12,
            11: 11,
            12: 14,
            13: 11,
        },
    )
    _set_print_layout(ws, landscape=True, fit_to_width=True, print_title_rows="1:3")
    _professionalize_sheet(
        ws,
        title_rows=1,
        freeze="A4",
        auto_filter=f"A3:{get_column_letter(last_col)}{gt_row}",
    )


def _short(s: str, m: int) -> str:
    s = (s or "").replace("\n", " ")
    return s if len(s) <= m else s[: m - 3] + "..."



def _instructions_weekly_43_line(sd: dict[str, Any]) -> str:
    """Match Summary tab: prefer ``mdhhs_weekly_minutes``, fall back to ``weekly_minutes``."""
    wk = int(sd.get("mdhhs_weekly_minutes") or sd.get("weekly_minutes") or 0)
    mo = int(sd.get("mdhhs_monthly_minutes") or 0)
    return f"Weekly minutes × 4.3 = monthly target ({wk} × 4.3 = {mo} min)."


def _instructions_apply_line_style(line: str) -> tuple[Font | None, PatternFill | None]:
    """Demonstration fonts/fills for the COLOR KEY section."""
    fn: Font | None = None
    fl: PatternFill | None = None
    low = line.lower()
    if "blue text" in low:
        fn = _f_input()
    elif "green text" in low:
        fn = _f_xref()
    elif "black text" in low:
        fn = _f_body()
    elif "yellow fill" in low:
        fl = FILL_TOTAL
    elif "orange fill" in low:
        fl = FILL_DEVIATION
    elif "peach fill" in low:
        fl = FILL_WEEKEND
    return fn, fl


# --- Tab 6: Instructions -----------------------------------------------------


def _build_instructions(ws: Worksheet, sd: dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    _title_row(ws, "Instructions — How to Read This Workbook", 1, 1, 4)

    daily = list(sd.get("daily_schedule") or [])
    catchup_iso = sd.get("catchup_date")
    shift_stats = _shift_type_stats(
        daily, str(catchup_iso) if catchup_iso else None
    )

    # Describe each shift type that actually appears in this plan, using
    # its aggregated stats — no hardcoded "Wed = Housework" narration.
    shift_lines: list[str] = []
    for shift_name, n, tot_min, _is_catch in shift_stats:
        avg = round_half_up(tot_min / n) if n > 0 else 0
        shift_lines.append(
            f"{shift_name} — {n} day(s), {tot_min:,} min total "
            f"(~{avg} min/day). Open the Daily Schedule tab to see each "
            f"day's placement."
        )
    if not shift_lines:
        shift_lines.append("No scheduled days in this plan.")

    sections: list[tuple[str, list[str]]] = [
        ("OVERVIEW", [
            "This workbook renders a calendar-specific home-care plan that "
            "reconciles exactly to the MDHHS-6064-P monthly authorization.",
            "Source of truth is the month-aware schedule — every day's clock "
            "times and task placements come from the calibrated daily list "
            "(ScheduleConfig.selected_weekdays ∪ selected_dates), not from "
            "a generic day-of-week rule.",
        ]),
        ("SHIFT TYPES (in this plan)", shift_lines),
        ("RECONCILIATION MATH", [
            _instructions_weekly_43_line(sd),
            "Every Daily Schedule row's task-minute columns sum to the "
            "Min (raw) column — if you edit a value, re-check that row.",
            "Task Reconciliation's grand-total variance reflects calendar shape — "
            "positive in 31-day months, negative in 28-day months. Billable always "
            "caps at the authorized amount per MDHHS ASM 144.",
        ]),
        ("COLOR KEY", [
            "Blue text — hardcoded MDHHS auth / task name inputs.",
            "Black text — formulas and computed values.",
            "Green text — cross-tab references.",
            "Yellow fill — totals row emphasis.",
            "Orange fill — deviation (catch-up) day.",
            "Peach fill — weekend rows.",
        ]),
    ]

    if catchup_iso:
        dt = _parse_iso(str(catchup_iso))
        month_upper = str(sd.get("month_name") or "").upper()
        date_upper = (
            f"{month_upper} {dt.day}".strip() if dt else str(catchup_iso).upper()
        )
        catch_entry = next(
            (d for d in daily if d.get("date") == str(catchup_iso)), {}
        )
        catch_tasks_list = [
            k for k, v in (catch_entry.get("tasks") or {}).items() if v
        ]
        catch_tasks = ", ".join(catch_tasks_list) or "no tasks"
        catch_dur = int(catch_entry.get("duration_min") or 0)
        clock_in = str(catch_entry.get("clock_in") or "")
        clock_out = str(catch_entry.get("clock_out") or "")
        weekday_text = dt.strftime("%A") if dt else ""
        normal_phrase = (
            f"instead of the normal {weekday_text} shift"
            if weekday_text
            else "as a calendar deviation"
        )

        sections.append(
            (
                f"{date_upper} DEVIATION",
                [
                    f"On {dt.strftime('%A, %B %d, %Y') if dt else catchup_iso} "
                    f"the caregiver works {clock_in}–{clock_out} "
                    f"({catch_dur} min) {normal_phrase}.",
                    (
                        f"The catch-up day performs: {catch_tasks}."
                        if catch_tasks_list
                        else "No tasks placed on the deviation day."
                    ),
                    "This closes the gap between calendar days in the month "
                    "and the 4.3-week MDHHS rule so the schedule reconciles "
                    "exactly.",
                ],
            )
        )

    row = 3
    dr = float(LAYOUT["data_row_height"])
    for title, bullets in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        sh = ws.cell(row, 1, value=title)
        sh.font = _f_section()
        sh.fill = FILL_SECTION
        sh.alignment = Alignment(horizontal="left", vertical="center")
        for cc in range(1, 5):
            ws.cell(row, cc).fill = FILL_SECTION
            ws.cell(row, cc).border = BORDER
        ws.row_dimensions[row].height = float(LAYOUT["section_row_height"])
        row += 1
        demo_color = title.strip().upper().startswith("COLOR KEY")
        for line in bullets:
            mark = ws.cell(row, 1, value="•")
            mark.font = _f_body()
            mark.alignment = Alignment(horizontal="center", vertical="top")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            body = ws.cell(row, 2, value=line)
            demo_fn, demo_fl = (
                _instructions_apply_line_style(line) if demo_color else (None, None)
            )
            body.font = demo_fn if demo_fn is not None else _f_body()
            body.fill = demo_fl if demo_fl is not None else FILL_NONE
            if demo_color and demo_fl is not None:
                for xc in range(2, 5):
                    ws.cell(row, xc).fill = demo_fl
            body.alignment = Alignment(wrap_text=True, vertical="top")
            _apply_border_range(ws, row, 1, row, 4)
            ws.row_dimensions[row].height = 36 if not demo_color else max(36.0, dr)
            row += 1
        _spacer_row(ws, row, height=6)
        row += 1

    _set_col_widths(ws, {1: 5, 2: 62, 3: 14, 4: 14})
    _set_print_layout(ws, landscape=True, fit_to_width=True, print_title_rows="1:2")
    _professionalize_sheet(ws, title_rows=1, freeze="A3")


def _finalize_workbook_polish(wb: Workbook) -> None:
    """Apply consistent tab styling and workbook metadata after every sheet is built."""
    wb.properties.title = "MDHHS Plan of Care Workbook"
    wb.properties.subject = "Home Help authorization schedule and reconciliation"
    wb.properties.creator = "MDHHS Plan Builder"
    for ws in wb.worksheets:
        if ws.title in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[ws.title]
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = ws.sheet_view.zoomScale or 90


# --- Public entry point ------------------------------------------------------


def build_xlsx(
    extracted_form: ExtractedForm,
    schedule: Any,
    validation_report: ValidationReport,
    output_path: str | Path,
    *,
    include_weekly_schedule: bool = True,
) -> None:
    """Render the full 7-tab workbook from a month-calibrated schedule.

    ``schedule`` may be a :class:`CalibratedSchedule` or its ``as_dict`` payload.
    Raises ``AssertionError`` if the Daily Schedule renderer drifts from the
    source schedule (an internal invariant, not a user-visible error).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sd = _as_schedule_dict(schedule)

    wb = Workbook()
    # openpyxl gives us one blank sheet; rename it to "Summary" instead of
    # creating an 8th "Sheet" tab.
    default_ws = wb.active
    assert default_ws is not None
    default_ws.title = "Summary"

    ws_summary = default_ws

    _build_summary(ws_summary, extracted_form, sd, validation_report)
    if include_weekly_schedule:
        build_weekly_schedule_sheet(
            extracted_form, schedule, wb, insert_at=1, validation_report=validation_report
        )

    ws_weekly = wb.create_sheet("Weekly Pattern")
    ws_daily = wb.create_sheet("Daily Schedule")
    ws_math = wb.create_sheet("Schedule Math")
    ws_recon = wb.create_sheet("Task Reconciliation")
    ws_instr = wb.create_sheet("Instructions")

    _build_weekly_pattern(ws_weekly, sd)
    first_r, last_r, dur_col, rt_col = _build_daily_schedule(
        ws_daily, extracted_form, sd
    )
    _build_schedule_math(ws_math, sd)
    _build_task_reconciliation(ws_recon, extracted_form, sd, validation_report)
    _build_instructions(ws_instr, sd)

    # Daily Schedule invariant — MUST hold for every row.
    _assert_daily_row_invariant(ws_daily, first_r, last_r, dur_col, rt_col)

    _finalize_workbook_polish(wb)
    wb.save(output_path)
    _try_libreoffice_recalc(output_path)

# --- LibreOffice headless recalc (optional, soft-fail) -----------------------


def _try_libreoffice_recalc(output_path: Path) -> None:
    exe = shutil.which("soffice") or shutil.which("libreoffice")
    if not exe:
        logger.warning(
            "LibreOffice (soffice) not installed — skipping formula recalc for %s",
            output_path.name,
        )
        return
    try:
        with tempfile.TemporaryDirectory() as td:
            subprocess.run(
                [
                    exe, "--calc", "--headless",
                    "--convert-to", "xlsx",
                    "--outdir", td,
                    str(output_path),
                ],
                check=False,
                timeout=30,
                capture_output=True,
            )
            recalc_path = Path(td) / output_path.name
            if recalc_path.exists():
                shutil.move(str(recalc_path), str(output_path))
    except (OSError, subprocess.SubprocessError) as e:
        logger.warning("LibreOffice recalc failed for %s: %s", output_path.name, e)
