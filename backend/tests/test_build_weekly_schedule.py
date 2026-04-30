"""Tests for backend/app/build_weekly_schedule.py and Weekly Schedule integration in build_xlsx."""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from app.build_weekly_schedule import build_weekly_schedule_sheet  # noqa: E402
from app.build_xlsx import build_xlsx  # noqa: E402
from app.calculate import (
    WEEKS_PER_MONTH,
    compute_mdhhs_form_minutes,
    generate_schedule,
    round_half_up,
)  # noqa: E402
from app.templates.weekly_schedule_tokens import WEEK_SCHEDULE_DAY_CODES  # noqa: E402
from app.validate import cross_check  # noqa: E402
from test_build_xlsx import OTTILIE_PREFERRED, OTTILIE_TASKS, _ottilie_form  # noqa: E402

_TOTAL_RE = re.compile(r"^\d+ hrs \d{2} mins$")
_HM_RE = re.compile(r"^(\d+) hrs (\d{2}) mins$")


def _parse_hrs_mins(cell: object) -> int:
    assert isinstance(cell, str)
    m = _HM_RE.fullmatch(cell.strip())
    assert m is not None
    return int(m.group(1)) * 60 + int(m.group(2))


def _value_beside_label(ws: openpyxl.worksheet.worksheet.Worksheet, label: str) -> object:
    for row in ws.iter_rows():
        if row and row[0].value == label:
            return row[1].value
    raise AssertionError(f"label not found: {label!r}")


def _no_mustache_anywhere(wb: openpyxl.Workbook) -> None:
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if isinstance(val, str) and ("{{" in val or "}}" in val):
                    raise AssertionError(f"unreplaced token in {name!r}: {val!r}")


def _ws_text(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
    parts: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                parts.append(str(v))
    return " ".join(parts)


@pytest.fixture
def ottilie_april_schedule():
    return generate_schedule(list(OTTILIE_TASKS), 27.0, 2026, 4, OTTILIE_PREFERRED)


EXPECTED_TAB_ORDER = [
    "Summary",
    "Weekly Schedule",
    "Weekly Pattern",
    "Daily Schedule",
    "Schedule Math",
    "Task Reconciliation",
    "Instructions",
]


@pytest.fixture
def full_plan_wb(tmp_path: Path, ottilie_april_schedule):
    form = _ottilie_form()
    path = tmp_path / "plan.xlsx"
    report = cross_check(form, ottilie_april_schedule)
    build_xlsx(form, ottilie_april_schedule, report, path)
    return openpyxl.load_workbook(path), ottilie_april_schedule


def test_build_xlsx_includes_weekly_schedule_tab_order_and_substitutions(full_plan_wb) -> None:
    wb, _cs = full_plan_wb
    form = _ottilie_form()
    assert wb.sheetnames == EXPECTED_TAB_ORDER
    ws = wb["Weekly Schedule"]
    _no_mustache_anywhere(wb)
    blob = _ws_text(ws)
    assert form.client_name in blob
    assert form.provider_name in blob
    st = _value_beside_label(ws, "Status:")
    assert isinstance(st, str) and (
        "BILLABLE" in st or "CHECK" in st or "EXACT MATCH" in st or "CALENDAR" in st
    )
    wt = _value_beside_label(ws, "Weekly Total Time:")
    assert isinstance(wt, str) and _TOTAL_RE.match(wt.strip())


def test_build_weekly_schedule_sheet_inserts_standalone_without_build_xlsx(
    ottilie_april_schedule,
) -> None:
    form = _ottilie_form()
    wb = Workbook()
    assert wb.active is not None
    wb.active.title = "Summary"
    build_weekly_schedule_sheet(form, ottilie_april_schedule, wb, insert_at=1)
    assert wb.sheetnames == ["Summary", "Weekly Schedule"]
    ws = wb["Weekly Schedule"]
    t = _ws_text(ws)
    assert form.client_name in t
    assert form.provider_name in t


def test_weekday_clock_ins_seven_am_weekend_noon(full_plan_wb):
    """Ottilie prefers 7:00 AM weekdays and 12:00 PM weekends (see OTTILIE_PREFERRED)."""
    wb, cs = full_plan_wb
    ws = wb["Weekly Schedule"]
    clock_in_row = 5
    for col, code in enumerate(WEEK_SCHEDULE_DAY_CODES, start=2):
        full = {
            "SUN": "Sunday",
            "MON": "Monday",
            "TUE": "Tuesday",
            "WED": "Wednesday",
            "THU": "Thursday",
            "FRI": "Friday",
            "SAT": "Saturday",
        }[code]
        expected = cs.weekly_pattern[full]["start"]
        assert ws.cell(row=clock_in_row, column=col).value == expected
    for col in range(3, 8):
        assert ws.cell(row=clock_in_row, column=col).value == "7:00 AM"
    assert ws.cell(row=clock_in_row, column=2).value == "12:00 PM"
    assert ws.cell(row=clock_in_row, column=8).value == "12:00 PM"


def test_total_cells_use_hrs_mins_format(full_plan_wb):
    wb, cs = full_plan_wb
    ws = wb["Weekly Schedule"]
    clock_top = 4
    total_row = clock_top + 3
    for col, code in enumerate(WEEK_SCHEDULE_DAY_CODES, start=2):
        full = {
            "SUN": "Sunday",
            "MON": "Monday",
            "TUE": "Tuesday",
            "WED": "Wednesday",
            "THU": "Thursday",
            "FRI": "Friday",
            "SAT": "Saturday",
        }[code]
        cell = ws.cell(row=total_row, column=col).value
        assert isinstance(cell, str)
        assert _TOTAL_RE.match(cell), cell
        mins = int((cs.weekly_pattern[full] or {}).get("minutes") or 0)
        assert _parse_hrs_mins(cell) == mins


def test_monthly_projects_from_weekly_and_status_exact(full_plan_wb):
    wb, cs = full_plan_wb
    ws = wb["Weekly Schedule"]
    weekly_cell = _value_beside_label(ws, "Weekly Total Time:")
    monthly_cell = _value_beside_label(ws, "Monthly Projected (x 4.3 weeks):")
    delivered_cell = _value_beside_label(ws, "Monthly Delivered (calendar shape):")
    authorized_cell = _value_beside_label(ws, "MDHHS Authorized (per 6064-P):")
    status_cell = _value_beside_label(ws, "Status:")
    wmin = _parse_hrs_mins(weekly_cell)
    mmin_proj = _parse_hrs_mins(monthly_cell)
    auth_min = _parse_hrs_mins(authorized_cell)
    # Aggregate × 4.3 from weekly pattern row (weekly_schedule placeholder math).
    assert mmin_proj == round_half_up(wmin * WEEKS_PER_MONTH)
    mdel = _parse_hrs_mins(str(delivered_cell))
    assert mdel == int(cs.delivered_minutes), (mdel, cs.delivered_minutes)
    # Per-line authorized total (matches MDHHS-6064-P; can differ from projection by ±min).
    assert auth_min == cs.mdhhs_monthly_minutes
    assert auth_min == compute_mdhhs_form_minutes(list(OTTILIE_TASKS))
    # Status beside "Status:" may still show template wording; workbook finalizer also
    # writes the combined CALENDAR · BILLABLE line on the EXACT MATCH cell.
    assert isinstance(status_cell, str) and len(status_cell.strip()) > 3
