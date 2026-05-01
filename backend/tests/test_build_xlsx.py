"""Tests for backend/app/build_xlsx.py.

Covers the seven-tab Ottilie April-2026 reference workbook and the regression
guard against the legacy _day_names_for_frequency placement helper.
"""

from __future__ import annotations

import ast
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from app.build_xlsx import build_xlsx  # noqa: E402
from app.calculate import (  # noqa: E402
    compute_mdhhs_form_amount,
    compute_monthly_minutes_rounded,
    default_config_for,
    generate_schedule,
)
from app.extract import ExtractedForm  # noqa: E402
from app.validate import cross_check  # noqa: E402


OTTILIE_TASKS = [
    {"task_name": "Bathing", "min_per_day": 16, "days_per_week": 7, "monthly_amount": 216.72},
    {"task_name": "Dressing", "min_per_day": 14, "days_per_week": 7, "monthly_amount": 189.63},
    {"task_name": "Grooming", "min_per_day": 8, "days_per_week": 7, "monthly_amount": 108.36},
    {"task_name": "Mobility", "min_per_day": 16, "days_per_week": 7, "monthly_amount": 216.72},
    {"task_name": "Toileting", "min_per_day": 6, "days_per_week": 7, "monthly_amount": 81.27},
    {"task_name": "Transferring", "min_per_day": 8, "days_per_week": 7, "monthly_amount": 108.36},
    {"task_name": "Medication", "min_per_day": 2, "days_per_week": 7, "monthly_amount": 27.09},
    {"task_name": "Meal Preparation", "min_per_day": 50, "days_per_week": 7, "monthly_amount": 677.25},
    {"task_name": "Housework", "min_per_day": 12, "days_per_week": 3, "monthly_amount": 69.66},
    {"task_name": "Laundry", "min_per_day": 21, "days_per_week": 2, "monthly_amount": 81.27},
    {"task_name": "Shopping for Food/Meds", "min_per_day": 15, "days_per_week": 2, "monthly_amount": 58.05},
    {"task_name": "Travel For Shopping", "min_per_day": 20, "days_per_week": 2, "monthly_amount": 77.40},
]

OTTILIE_PREFERRED = {"weekday_start": "7:00 AM", "weekend_start": "12:00 PM"}


def _ottilie_form() -> ExtractedForm:
    return ExtractedForm(
        client_name="Ottilie Smith",
        client_id="80738972",
        county_name="82-WAYNE",
        case_number="350195-2",
        asw_name="K Abdelkhaliq",
        asw_phone="313-407-4457",
        asw_email="AbdelkhaliqK@michigan.gov",
        auth_date="04/01/2026",
        provider_name="Alegria Home Care",
        pay_rate=27.00,
        tasks=list(OTTILIE_TASKS),
        monthly_total_time_str="70:48",
        monthly_total_amount=1911.78,
    )


@pytest.fixture(scope="module")
def built_workbook(tmp_path_factory: pytest.TempPathFactory) -> tuple[Path, "openpyxl.Workbook"]:
    form = _ottilie_form()
    cs = generate_schedule(list(OTTILIE_TASKS), 27.0, 2026, 4, OTTILIE_PREFERRED)
    report = cross_check(form, cs)
    out_dir = tmp_path_factory.mktemp("xlsx_out")
    path = out_dir / "ottilie_april_2026.xlsx"
    build_xlsx(form, cs.as_dict(), report, path)
    wb = openpyxl.load_workbook(path)
    return path, wb


def test_seven_tabs_present(built_workbook: tuple[Path, "openpyxl.Workbook"]) -> None:
    _, wb = built_workbook
    assert wb.sheetnames == [
        "Summary",
        "Weekly Schedule",
        "Weekly Pattern",
        "Daily Schedule",
        "Schedule Math",
        "Task Reconciliation",
        "Instructions",
    ]


def test_summary_client_and_asw_fields_from_extracted_form(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    """Summary tab maps ExtractedForm case metadata (regression: rebuild must persist these)."""
    _, wb = built_workbook
    ws = wb["Summary"]
    want = _ottilie_form()

    def left_val(label: str) -> object:
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == label:
                return ws.cell(r, 3).value
        raise AssertionError(f"missing label {label!r}")

    def asw_val(label: str) -> object:
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 4).value == label:
                return ws.cell(r, 6).value
        raise AssertionError(f"missing ASW label {label!r}")

    assert left_val("Client ID") == want.client_id
    assert left_val("County") == want.county_name
    assert left_val("Case Number") == want.case_number
    assert left_val("Provider Agency") == want.provider_name
    assert left_val("Auth Date") == want.auth_date
    assert asw_val("ASW Name") == want.asw_name


def test_schedule_math_cumulative_header_names_form_dollar_cap(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    _, wb = built_workbook
    ws = wb["Schedule Math"]
    hits: list[str] = []
    for row in ws.iter_rows(values_only=False):
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("DAILY CUMULATIVE"):
                hits.append(v)
    assert len(hits) == 1
    assert "billable capped at $1,911.78" in hits[0]


def test_instructions_reconciliation_lists_weekly_and_monthly_derived_totals(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    """Instructions tab interpolates weekly minutes (988) and mdhhs monthly minutes (4248)."""
    _, wb = built_workbook
    ws = wb["Instructions"]
    found = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "Weekly minutes × 4.3" in cell:
                found = True
                assert "(988 × 4.3 = 4248 min)" in cell
    assert found


def test_daily_schedule_row_totals_match_duration(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    _, wb = built_workbook
    ws = wb["Daily Schedule"]

    header_row = 3
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    assert "Min (raw)" in headers and "Row Total" in headers
    dur_col = headers.index("Min (raw)") + 1
    rt_col = headers.index("Row Total") + 1

    day_rows = 0
    week_band_rows = 0
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if not first or str(first).startswith("MONTHLY TOTAL"):
            continue
        is_week_band = (
            isinstance(first, str) and first.startswith("Week of ")
            and "Subtotal" in first
        )
        duration = ws.cell(r, dur_col).value
        row_total = ws.cell(r, rt_col).value
        if duration is None or isinstance(duration, bool):
            continue
        if not isinstance(duration, (int, float)):
            continue
        task_sum = 0
        for c in range(dur_col + 1, rt_col):
            v = ws.cell(r, c).value
            # Skip non-numeric cells (e.g. ``HH:MM`` display column)
            if v is None or not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            task_sum += int(v)
        assert int(task_sum) == int(duration), (
            f"Row {r}: Σ(tasks)={task_sum} != duration={duration}"
        )
        assert int(row_total) == int(duration), (
            f"Row {r}: row_total={row_total} != duration={duration}"
        )
        if is_week_band:
            week_band_rows += 1
        else:
            day_rows += 1
    assert day_rows == 30, f"Expected 30 schedule day-rows, walked {day_rows}"
    assert week_band_rows >= 1, "Expected at least one weekly subtotal band row"


def test_catchup_day_present_april_2026(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    _, wb = built_workbook
    ws = wb["Daily Schedule"]

    catch_row = None
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "2026-04-29":
            catch_row = r
            break
    assert catch_row is not None, "No row for 2026-04-29 in Daily Schedule"

    shift_type = ws.cell(catch_row, 3).value
    clock_in = ws.cell(catch_row, 4).value
    clock_out = ws.cell(catch_row, 5).value
    duration = ws.cell(catch_row, 6).value

    assert str(shift_type) == "CATCHUP"
    assert int(duration) == 176

    def _hm(v: object) -> str:
        if hasattr(v, "strftime"):
            return v.strftime("%I:%M %p").lstrip("0")  # type: ignore[union-attr]
        return str(v).strip()

    assert _hm(clock_in) == "7:00 AM"
    assert _hm(clock_out) == "9:56 AM"


def test_grand_total_equals_4248_per_line_minutes(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    path, _ = built_workbook
    wb = openpyxl.load_workbook(path, data_only=True)
    target_tabs = ("Summary", "Schedule Math")
    for name in target_tabs:
        ws = wb[name]
        found = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)) and int(cell) == 4248 and not isinstance(cell, bool):
                    found = True
                    break
            if found:
                break
        assert found, f"Monthly total 4248 not found on tab {name!r}"


def test_scheduled_dollar_is_1911_78(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    path, _ = built_workbook
    wb = openpyxl.load_workbook(path, data_only=True)
    target_tabs = ("Summary", "Schedule Math")
    for name in target_tabs:
        ws = wb[name]
        found = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                    if abs(float(cell) - 1911.78) < 0.005:
                        found = True
                        break
            if found:
                break
        assert found, f"$1,911.78 not found on tab {name!r}"


def test_task_reconciliation_badge_reflects_validation(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    """Task Reconciliation's status badge mirrors ``ValidationReport.validation_status``.

    Ottilie April 2026 is an all-checks-pass case with billing equality
    (``BILLABLE_EXACT``). The XLSX shows the human-readable badge next to
    the TOTAL row. The tab must still surface a Variance column for reviewers.
    """
    form = _ottilie_form()
    cs = generate_schedule(list(OTTILIE_TASKS), 27.0, 2026, 4, OTTILIE_PREFERRED)
    report = cross_check(form, cs)

    badge_labels = {
        "BILLABLE_EXACT": "BILLABLE EXACT ✓",
        "BILLABLE_AT_CAP": "BILLABLE AT CAP ✓",
        "BILLABLE_UNDER_CAP": "BILLABLE UNDER CAP ✓",
        "INVALID": "INVALID ✗",
    }
    expected = badge_labels.get(report.validation_status, report.validation_status)

    _, wb = built_workbook
    ws = wb["Task Reconciliation"]

    header_row = 3
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    assert "Variance" in headers, f"'Variance' column missing: {headers!r}"

    total_row = None
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "TOTAL":
            total_row = r
            break
    assert total_row is not None, "No TOTAL row on Task Reconciliation tab"

    pct_share_col = headers.index("% Share") + 1
    assert ws.cell(total_row, pct_share_col).value == pytest.approx(1.0)

    status_text = None
    for r in range(total_row + 1, min(total_row + 4, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "BILLABLE" in v:
                status_text = v
                break
        if status_text:
            break
    assert status_text == expected, (
        f"Expected {expected!r} near TOTAL row (validation_status={report.validation_status!r}), "
        f"got {status_text!r}"
    )


def test_task_reconciliation_has_occurrence_column(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    """Task Reconciliation must surface the per-task session count from
    ``sd['task_occurrence_counts']`` so the workbook doesn't re-derive
    occurrences from a frequency table.
    """
    _, wb = built_workbook
    ws = wb["Task Reconciliation"]

    header_row = 3
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    assert "Occurrences" in headers, (
        f"'Occurrences' column missing from Task Reconciliation: {headers!r}"
    )
    assert "% Share" in headers, f"Expected trailing % Share column: {headers!r}"
    task_col = headers.index("Task") + 1
    occ_col = headers.index("Occurrences") + 1

    # Walk data rows until Subtotal / TOTAL; every task must have a non-negative int.
    seen_any = False
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, task_col).value
        if not name:
            break
        sname = str(name)
        if sname.startswith("Subtotal"):
            break
        if sname == "TOTAL":
            break
        occ = ws.cell(r, occ_col).value
        assert isinstance(occ, int) and occ >= 0, (
            f"Row {r} ({name!r}) has non-integer occurrences: {occ!r}"
        )
        seen_any = True
    assert seen_any, "Task Reconciliation has no data rows"


def test_workbook_monthly_dollar_is_single_value(tmp_path: Path) -> None:
    """Summary has one Monthly $ row aligned with mdhhs_form_amount — no schedule duplicate."""
    tasks = [
        {
            "task_name": "Bathing",
            "min_per_day": 15,
            "days_per_week": 7,
            "monthly_amount": 0.0,
        },
        {
            "task_name": "Mobility",
            "min_per_day": 15,
            "days_per_week": 5,
            "monthly_amount": 0.0,
        },
    ]
    pay = 27.0
    form = ExtractedForm(
        client_name="Single $ Row",
        pay_rate=pay,
        tasks=list(tasks),
        monthly_total_amount=compute_mdhhs_form_amount(tasks, pay),
    )
    cs = generate_schedule(tasks, pay, 2026, 4)
    report = cross_check(form, cs)
    path = tmp_path / "one_monthly_dollar_row.xlsx"
    build_xlsx(form, cs.as_dict(), report, path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    labels = [
        ws.cell(r, 1).value
        for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, 1).value, str)
    ]
    assert "Monthly $ (schedule)" not in labels
    assert "Monthly $ (form line-sum)" not in labels
    assert sum(1 for x in labels if x == "Monthly $") == 1
    expected = float(cs.as_dict().get("mdhhs_form_amount") or 0.0)
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Monthly $":
            cell = ws.cell(r, 3).value
            assert cell is not None
            assert abs(float(cell) - expected) < 0.005
            break
    else:
        raise AssertionError("Monthly $ row missing on Summary tab")


def test_summary_has_hours_minutes_companion(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    """Summary tab exposes ``Weekly HH:MM`` / ``Monthly HH:MM`` next to minute rows.

    For Ottilie's April 2026 plan the expected HH:MM values are:
    - Weekly: 988 min == ``16:28``
    - Monthly: 4,248 min == ``70:48``
    """
    _, wb = built_workbook
    ws = wb["Summary"]

    label_cells: dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str):
            continue
        value = ws.cell(r, 3).value
        if isinstance(value, str):
            label_cells[label] = value
            continue
        value2 = ws.cell(r, 2).value
        if isinstance(value2, str):
            label_cells[label] = value2

    assert "Weekly HH:MM" in label_cells, (
        f"Summary is missing 'Weekly HH:MM' row: {list(label_cells)!r}"
    )
    assert label_cells["Weekly HH:MM"] == "16:28", (
        f"Weekly HH:MM mismatch: {label_cells['Weekly HH:MM']!r}"
    )

    assert "Monthly HH:MM" in label_cells, (
        f"Summary is missing 'Monthly HH:MM' row: {list(label_cells)!r}"
    )
    assert label_cells["Monthly HH:MM"] == "70:48", (
        f"Monthly HH:MM mismatch: {label_cells['Monthly HH:MM']!r}"
    )


def test_weekly_pattern_has_professional_total_row(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    _, wb = built_workbook
    ws = wb["Weekly Pattern"]

    assert ws.cell(11, 1).value == "WEEKLY TOTAL"
    assert ws.cell(11, 4).value == 988
    assert ws.cell(11, 5).value == "16:28 (16h 28m)"
    assert ws.cell(11, 6).value == "16 hrs 28 mins"
    assert ws.freeze_panes == "A4"
    assert ws.auto_filter.ref == "A3:F11"


def test_workbook_has_professional_tab_colors_and_tables(
    built_workbook: tuple[Path, "openpyxl.Workbook"],
) -> None:
    _, wb = built_workbook
    assert wb["Summary"].sheet_properties.tabColor is not None
    assert "WeeklyPatternTable" in wb["Weekly Pattern"].tables
    assert "TaskReconciliationTable" in wb["Task Reconciliation"].tables


def test_task_reconciliation_ehhs_and_travel_subtotals(
    tmp_path: Path,
) -> None:
    """Subtotal rows sum non-travel vs travel authorization like validate EHHS."""
    tasks = [
        {
            "task_name": "Bathing",
            "min_per_day": 30,
            "days_per_week": 7,
            "monthly_amount": 0.0,
        },
        {
            "task_name": "Travel For Shopping",
            "min_per_day": 20,
            "days_per_week": 2,
            "monthly_amount": 0.0,
        },
        {
            "task_name": "Travel For Laundry",
            "min_per_day": 15,
            "days_per_week": 1,
            "monthly_amount": 0.0,
        },
    ]
    form = ExtractedForm(
        client_name="EHHS Mix",
        pay_rate=18.0,
        tasks=list(tasks),
        monthly_total_amount=0.0,
    )
    y, m = 2026, 6
    pw = {"weekday_start": "7:00 AM", "weekend_start": "12:00 PM"}
    cfg = default_config_for(tasks, y, m, pw)
    cs = generate_schedule(tasks, 18.0, y, m, pw, config=cfg)
    report = cross_check(form, cs)
    path = tmp_path / "ehhs_travel_subtotals.xlsx"
    build_xlsx(form, cs.as_dict(), report, path)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Task Reconciliation"]

    want_ehhs = compute_monthly_minutes_rounded(30, 7)
    want_travel = compute_monthly_minutes_rounded(20, 2) + compute_monthly_minutes_rounded(
        15, 1
    )

    ehhs_row = tr_row = None
    for r in range(4, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label:
            continue
        ls = str(label)
        if "Personal Care + IADL + Complex Care" in ls:
            ehhs_row = r
        if "Travel (excluded from 179.59-hr EHHS cap)" in ls:
            tr_row = r
    assert ehhs_row is not None, "EHHS subtotal row missing"
    assert tr_row is not None, "Travel subtotal row missing"
    # Columns: Task | Min/Day | Day HH:MM | Days/Wk | Auth Min (…) | … — auth is col 5
    auth_min_col = 5
    assert ws.cell(ehhs_row, auth_min_col).value == want_ehhs
    assert ws.cell(tr_row, auth_min_col).value == want_travel


def test_no_legacy_placement_helper_used() -> None:
    """Regression guard (Bug 3): build_xlsx.py must not import or call
    ``_day_names_for_frequency`` — the month-aware schedule is the only
    source of truth for day-by-day task placements."""
    src_path = Path(__file__).resolve().parents[1] / "app" / "build_xlsx.py"
    src = src_path.read_text(encoding="utf-8")

    assert "_day_names_for_frequency" not in src, (
        "build_xlsx.py still references _day_names_for_frequency — the legacy "
        "task-placement helper must not be used. Read from "
        'schedule["daily_schedule"] instead.'
    )

    tree = ast.parse(src)
    for node in ast.walk(tree):
        if isinstance(node, ast.ImportFrom):
            for alias in node.names:
                assert alias.name != "_day_names_for_frequency", (
                    f"Illegal import at line {node.lineno}: "
                    "_day_names_for_frequency is a legacy helper."
                )
        if isinstance(node, ast.Call):
            func = node.func
            if isinstance(func, ast.Name) and func.id == "_day_names_for_frequency":
                raise AssertionError(
                    f"Illegal call at line {node.lineno}: _day_names_for_frequency"
                )
            if isinstance(func, ast.Attribute) and func.attr == "_day_names_for_frequency":
                raise AssertionError(
                    f"Illegal call at line {node.lineno}: _day_names_for_frequency"
                )
