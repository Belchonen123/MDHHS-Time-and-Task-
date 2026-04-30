"""End-to-end: every place that surfaces monthly $ for one plan must agree, to the cent.

Regression net for the dual-path (aggregate vs per-line) monthly $ / minutes bug.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from app.build_xlsx import build_xlsx
from app.calculate import (
    compute_mdhhs_form_amount,
    compute_mdhhs_form_minutes,
    generate_schedule,
)
from app.extract import ExtractedForm
from app.validate import cross_check

TASKS = [
    {"task_name": "Bathing", "min_per_day": 15, "days_per_week": 7, "monthly_amount": 0.0},
    {"task_name": "Dressing", "min_per_day": 10, "days_per_week": 7, "monthly_amount": 0.0},
    {"task_name": "Mobility", "min_per_day": 15, "days_per_week": 5, "monthly_amount": 0.0},
    {"task_name": "Housework", "min_per_day": 30, "days_per_week": 3, "monthly_amount": 0.0},
    {"task_name": "Laundry", "min_per_day": 45, "days_per_week": 2, "monthly_amount": 0.0},
]
PAY = 27.00


def _sample_form() -> ExtractedForm:
    amt = compute_mdhhs_form_amount(TASKS, PAY)
    return ExtractedForm(
        client_name="Test",
        client_id="1",
        pay_rate=PAY,
        tasks=list(TASKS),
        monthly_total_amount=amt,
    )


def test_calculate_monthly_amount_equals_form_amount():
    cs = generate_schedule(TASKS, PAY, 2026, 4)
    assert cs.mdhhs_monthly_amount == cs.mdhhs_form_amount
    assert cs.mdhhs_monthly_minutes == compute_mdhhs_form_minutes(TASKS)


def test_workbook_monthly_dollar_matches_calculate(tmp_path: Path):
    """Summary `Monthly $` row and Schedule Math `Monthly $ target` match calculate."""
    form = _sample_form()
    cs = generate_schedule(TASKS, PAY, 2026, 4)
    sd = cs.as_dict()
    report = cross_check(form, cs)
    out = tmp_path / "test.xlsx"
    build_xlsx(form, sd, report, out)

    wb = openpyxl.load_workbook(out, data_only=True)
    expected = float(cs.mdhhs_form_amount)

    found_exact: list[float] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value != "Monthly $":
                continue
            for cc in range(2, ws.max_column + 1):
                v = ws.cell(r, cc).value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    found_exact.append(float(v))
                    break

    assert found_exact, "no 'Monthly $' label found in workbook"
    for v in found_exact:
        assert v == expected, (
            f"workbook shows ${v}, calculate gives ${expected}"
        )

    sm = wb["Schedule Math"]
    for r in range(1, sm.max_row + 1):
        if sm.cell(r, 1).value == "Monthly $ target":
            v = sm.cell(r, 4).value
            assert v is not None
            assert float(v) == expected
            break
    else:
        raise AssertionError("Schedule Math: 'Monthly $ target' row missing")


def test_check_6_passes_with_no_inflated_tolerance():
    form = _sample_form()
    cs = generate_schedule(TASKS, PAY, 2026, 4)
    report = cross_check(form, cs.as_dict())
    check_6 = next(c for c in report.checks if c.number == 6)
    assert check_6.passed, f"Check 6 should pass cleanly: {check_6.detail}"
    tol = str(check_6.tolerance)
    assert "$0.02" in tol
    assert "one-session" not in tol


def test_per_task_form_total_minutes_matches_schedule_target():
    """``compute_mdhhs_form_minutes`` (Σ unrounded, round once) equals schedule target."""
    cs = generate_schedule(TASKS, PAY, 2026, 4)
    form_total = compute_mdhhs_form_minutes(TASKS)
    assert form_total == cs.mdhhs_monthly_minutes
