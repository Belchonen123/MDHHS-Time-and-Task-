"""End-to-end weekly schedule workbook checks across calendar month shapes.

Catches template drift and ensures the Weekly Schedule sheet (embedded in the
full workbook via ``build_xlsx``) stays aligned with ``generate_schedule``.
"""

from __future__ import annotations

import calendar
import re
import sys
from datetime import date
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from app.build_xlsx import build_xlsx  # noqa: E402
from app.calculate import ScheduleConfig, TaskPlacement, generate_schedule  # noqa: E402
from app.templates.weekly_schedule_tokens import WEEK_SCHEDULE_DAY_CODES  # noqa: E402
from app.validate import cross_check  # noqa: E402
from test_build_xlsx import OTTILIE_PREFERRED, OTTILIE_TASKS, _ottilie_form  # noqa: E402

_FULL_WEEK = (
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
)

_HM_RE = re.compile(r"^(\d+) hrs (\d{2}) mins$")
_AMPM_RE = re.compile(r"^(\d{1,2}):(\d{2})\s*([AP]M)$", re.IGNORECASE)

# Template layout (must match ``weekly_schedule_tokens.write_weekly_schedule_template``).
_CLOCK_TOP = 4
_CLOCK_IN_ROW = _CLOCK_TOP + 1
_CLOCK_OUT_ROW = _CLOCK_TOP + 2
_TOTAL_ROW = _CLOCK_TOP + 3

# Monthly *projected* row (weekly Σ × 4.3) for Ottilie weekly pattern — not per-line auth.
_MONTHLY_PROJECTED_REF_MIN = 4248


def _count_weekday_in_month(year: int, month: int, weekday: int) -> int:
    """``weekday``: ``date.weekday()`` (Mon=0 … Sun=6)."""
    _, nd = calendar.monthrange(year, month)
    return sum(
        1 for d in range(1, nd + 1) if date(year, month, d).weekday() == weekday
    )


def _ottilie_custom_config() -> ScheduleConfig:
    """Housework Tue/Wed/Thu; Laundry + Shopping + Travel Tue/Thu only."""
    placements: list[TaskPlacement] = []
    for t in OTTILIE_TASKS:
        name = str(t["task_name"])
        mpd = int(t["min_per_day"])
        dpw = int(t["days_per_week"])
        if name == "Housework":
            sw = ["Tuesday", "Wednesday", "Thursday"]
        elif name in ("Laundry", "Shopping for Food/Meds", "Travel For Shopping"):
            sw = ["Tuesday", "Thursday"]
        elif dpw == 7:
            sw = list(_FULL_WEEK)
        else:
            raise AssertionError(f"Unexpected task in OTTILIE_TASKS: {name!r} dpw={dpw}")
        placements.append(
            TaskPlacement(
                task_name=name,
                min_per_day=mpd,
                days_per_week=dpw,
                selected_weekdays=sw,
            )
        )
    return ScheduleConfig(tasks=placements)


def _parse_hrs_mins(cell: object) -> int:
    assert isinstance(cell, str)
    m = _HM_RE.fullmatch(cell.strip())
    assert m is not None, cell
    return int(m.group(1)) * 60 + int(m.group(2))


def _ampm_to_minutes(s: str) -> int:
    m = _AMPM_RE.match((s or "").strip())
    assert m is not None, s
    h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    if not (1 <= h <= 12 and 0 <= mi <= 59):
        raise ValueError(s)
    if ap == "AM":
        h24 = 0 if h == 12 else h
    else:
        h24 = 12 if h == 12 else h + 12
    return h24 * 60 + mi


def _duration_minutes_same_day(clock_in: str, clock_out: str) -> int:
    a = _ampm_to_minutes(clock_in)
    b = _ampm_to_minutes(clock_out)
    d = b - a
    if d < 0:
        d += 24 * 60
    return d


def _value_beside_label(ws: object, label: str) -> object:
    for row in ws.iter_rows():
        if row and row[0].value == label:
            return row[1].value
    raise AssertionError(f"label not found: {label!r}")


def _parse_variance_mins(cell: object) -> int:
    assert isinstance(cell, str)
    s = cell.strip()
    if s == "0 mins":
        return 0
    if s.startswith("-"):
        return -int(s[1:].replace(" mins", "").strip())
    return int(s.replace(" mins", "").strip())


def _code_to_full() -> dict[str, str]:
    return {
        "SUN": "Sunday",
        "MON": "Monday",
        "TUE": "Tuesday",
        "WED": "Wednesday",
        "THU": "Thursday",
        "FRI": "Friday",
        "SAT": "Saturday",
    }


@pytest.mark.parametrize(
    "year,month",
    [
        (2026, 4),  # 30-day month; April 2026 has 5 Wednesdays
        (2026, 7),  # 31-day month; July 2026 has 4 Saturdays
        (2026, 2),  # 28-day February
    ],
    ids=["apr2026_30d_5wed", "jul2026_31d_4sat", "feb2026_28d"],
)
def test_weekly_schedule_e2e_ottilie_custom_placement(
    tmp_path: Path, year: int, month: int
) -> None:
    label = f"{year}-{month:02d}"
    nd = calendar.monthrange(year, month)[1]
    if month == 4:
        assert nd == 30, label
        assert _count_weekday_in_month(year, month, 2) == 5, label  # Wednesday
    elif month == 7:
        assert nd == 31, label
        assert _count_weekday_in_month(year, month, 5) == 4, label  # Saturday
    elif month == 2:
        assert nd == 28, label

    cfg = _ottilie_custom_config()
    cal = generate_schedule(
        list(OTTILIE_TASKS),
        27.0,
        year,
        month,
        OTTILIE_PREFERRED,
        config=cfg,
    )
    wp = cal.weekly_pattern

    assert wp["Tuesday"]["minutes"] == 188, label
    assert wp["Thursday"]["minutes"] == 188, label
    assert wp["Wednesday"]["minutes"] == 132, label
    for d in ("Monday", "Friday", "Saturday", "Sunday"):
        assert wp[d]["minutes"] == 120, f"{label} {d}"

    out = tmp_path / f"weekly_e2e_{label}.xlsx"
    form = _ottilie_form()
    report = cross_check(form, cal)
    build_xlsx(form, cal, report, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Weekly Schedule"]
    code_to_full = _code_to_full()

    for col, code in enumerate(WEEK_SCHEDULE_DAY_CODES, start=2):
        full = code_to_full[code]
        cin = ws.cell(row=_CLOCK_IN_ROW, column=col).value
        cout = ws.cell(row=_CLOCK_OUT_ROW, column=col).value
        total_cell = ws.cell(row=_TOTAL_ROW, column=col).value
        assert isinstance(cin, str) and isinstance(cout, str), (label, full)
        assert isinstance(total_cell, str), (label, full)
        exp_min = int(wp[full]["minutes"])
        assert _parse_hrs_mins(total_cell) == exp_min, (label, full, total_cell)
        assert _duration_minutes_same_day(cin, cout) == exp_min, (
            label,
            full,
            cin,
            cout,
            exp_min,
        )

    monthly_cell = _value_beside_label(ws, "Monthly Projected (x 4.3 weeks):")
    monthly_mins = _parse_hrs_mins(monthly_cell)
    assert abs(monthly_mins - _MONTHLY_PROJECTED_REF_MIN) <= 1, (
        f"{label}: monthly projected {monthly_mins} not within ±1 of {_MONTHLY_PROJECTED_REF_MIN}"
    )

    auth_cell = _value_beside_label(ws, "MDHHS Authorized (per 6064-P):")
    var_cell = _value_beside_label(ws, "Variance:")
    status_cell = str(_value_beside_label(ws, "Status:"))

    auth_mins = _parse_hrs_mins(auth_cell)
    variance_parsed = _parse_variance_mins(var_cell)
    assert variance_parsed == monthly_mins - auth_mins, (
        f"{label}: variance cell {variance_parsed!r} != monthly−auth "
        f"{monthly_mins - auth_mins}"
    )
    assert "BILLABLE" in status_cell or "INVALID" in status_cell, (
        f"{label}: expected billing badge or calendar note in status ({status_cell!r})"
    )
