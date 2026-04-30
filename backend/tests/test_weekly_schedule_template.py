"""Regression tests for the client-facing weekly schedule Excel template."""

from __future__ import annotations

import sys
from dataclasses import fields
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from app.templates.weekly_schedule_tokens import (  # noqa: E402
    WEEKLY_SCHEDULE_TOKENS,
    WEEK_SCHEDULE_DAY_CODES,
)


def _template_path() -> Path:
    return (
        Path(__file__).resolve().parents[1]
        / "app"
        / "templates"
        / "weekly_schedule_template.xlsx"
    )


def _iter_sheet_text(wb: openpyxl.Workbook, title: str) -> str:
    ws = wb[title]
    parts: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                parts.append(str(val))
    return "\n".join(parts)


@pytest.fixture(scope="module")
def weekly_wb() -> openpyxl.Workbook:
    path = _template_path()
    assert path.is_file(), f"missing template: {path}"
    return openpyxl.load_workbook(path)


def test_weekly_schedule_sheet_exists(weekly_wb: openpyxl.Workbook) -> None:
    assert "Weekly Schedule" in weekly_wb.sheetnames


def test_field_guide_sheet_exists(weekly_wb: openpyxl.Workbook) -> None:
    assert "Field Guide" in weekly_wb.sheetnames


def test_all_task_placeholders_in_schedule_sheet(weekly_wb: openpyxl.Workbook) -> None:
    blob = _iter_sheet_text(weekly_wb, "Weekly Schedule")
    for day in WEEK_SCHEDULE_DAY_CODES:
        for i in range(1, 13):
            ph = getattr(WEEKLY_SCHEDULE_TOKENS, f"TASK_{day}_{i}")
            assert ph in blob, f"missing task placeholder {ph}"


def test_header_clock_and_summary_placeholders(weekly_wb: openpyxl.Workbook) -> None:
    blob = _iter_sheet_text(weekly_wb, "Weekly Schedule")
    for name in ("CLIENT_NAME", "SHIFT_START", "PROVIDER_NAME"):
        ph = getattr(WEEKLY_SCHEDULE_TOKENS, name)
        assert ph in blob, f"missing header token {ph}"
    for day in WEEK_SCHEDULE_DAY_CODES:
        for prefix in ("CLOCK_IN", "CLOCK_OUT", "TOTAL"):
            ph = getattr(WEEKLY_SCHEDULE_TOKENS, f"{prefix}_{day}")
            assert ph in blob, f"missing clock token {ph}"
    for name in (
        "WEEKLY_TOTAL",
        "MONTHLY_TOTAL",
        "AUTHORIZED_TOTAL",
        "VARIANCE",
        "STATUS",
    ):
        ph = getattr(WEEKLY_SCHEDULE_TOKENS, name)
        assert ph in blob, f"missing summary token {ph}"


def test_field_guide_lists_every_token(weekly_wb: openpyxl.Workbook) -> None:
    ws = weekly_wb["Field Guide"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    expected = {f.name for f in fields(WEEKLY_SCHEDULE_TOKENS)}
    found = {r[0] for r in rows if r and r[0] is not None}
    assert found == expected
